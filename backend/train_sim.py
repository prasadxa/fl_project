"""
Standalone FL Simulation — train_sim.py
========================================
Runs Federated Averaging (FedAvg) in a single Python process.

No network.  No gRPC.  No Ray.  No server/client processes.
The FedAvg math is identical to the real server.py + client.py stack:
  each client trains locally, then the server performs a sample-weighted
  average of all client weight tensors.

Checkpoint behaviour (mirrors server.py):
  - On startup, loads models/global_model.pth if it exists (resume by default).
  - After every round saves:
      models/global_model.pth                          ← always overwritten
      models/global_model_round_N.pth                  ← clean per-round file
      models/global_model_round_N_YYYYMMDD_HHMMSS.pth  ← timestamped archive

Usage (run from project root):
    python src/train_sim.py                        # 5 rounds, 10 epochs, resume from checkpoint
    python src/train_sim.py --rounds 10            # 10 rounds
    python src/train_sim.py --rounds 5 --fresh     # ignore checkpoint, random init
    python src/train_sim.py --rounds 5 --eval      # evaluate on test set every round
    python src/train_sim.py --rounds 3 --epochs 10 # 10 local epochs per client (default)
"""

from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import torch
from torch.utils.data import DataLoader

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

# -- Force UTF-8 output on Windows (avoids cp1252 UnicodeEncodeError) ---------
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# -- Project imports ----------------------------------------------------------
sys.path.insert(0, str(Path(__file__).parent))

from dataset import (
    CLASS_NAMES,
    NUM_CLASSES,
    compute_class_weights,
    get_client_loader,
    get_test_loader,
)
from model import (
    MedicalCNN,
    evaluate,
    get_parameters,
    set_parameters,
    train_fedprox,
)

# -- Constants ----------------------------------------------------------------
PROJ_ROOT = Path(__file__).parent.parent
MODELS_DIR = PROJ_ROOT / "models"
MODELS_DIR.mkdir(exist_ok=True)

NUM_CLIENTS = 3


# Auto-select best available device: MPS (Apple Silicon) > CUDA > CPU
def _best_device() -> str:
    if torch.backends.mps.is_available():
        return "mps"
    if torch.cuda.is_available():
        return "cuda"
    return "cpu"


def _try_compile(model: "MedicalCNN") -> "MedicalCNN":
    """
    Wrap model with torch.compile if available.
    Uses the plain 'eager' backend which is fully compatible with MPS and
    avoids max_pool2d_backward warnings triggered by 'aot_eager' on Apple Silicon.
    Falls back silently if compilation is not supported.
    """
    try:
        return torch.compile(model, backend="eager")  # type: ignore[return-value]
    except Exception:
        return model


DEVICE = _best_device()

DEFAULT_ROUNDS = 30
DEFAULT_EPOCHS = 2
DEFAULT_BATCH = 64  # doubled — MPS/CUDA can handle larger batches
DEFAULT_LR = 0.01
DEFAULT_MU = 0.01
# On macOS MPS, multiprocessing DataLoader workers are 15-30x slower due to
# spawn overhead. Use 0 (in-process) for MPS/CPU, only use workers for CUDA.
DEFAULT_WORKERS = 0
NUM_EVAL_WORKERS = 0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _ts() -> str:
    """UTC timestamp string for log lines."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _banner(text: str, width: int = 65, char: str = "=") -> None:
    print(char * width)
    print(f"  {text}")
    print(char * width)


def _hline(width: int = 65) -> None:
    print("-" * width)


def _acc_bar(acc: float, width: int = 30) -> str:
    """Return a coloured ASCII bar like  [████████████░░░░░░░░░░░░░░░░░░]  72.3%"""
    filled = int(round(acc * width))
    bar = "█" * filled + "░" * (width - filled)
    pct = acc * 100
    # Simple ANSI colour: green ≥80%, yellow ≥55%, red otherwise
    if pct >= 80:
        colour = "\033[92m"  # bright green
    elif pct >= 55:
        colour = "\033[93m"  # bright yellow
    else:
        colour = "\033[91m"  # bright red
    reset = "\033[0m"
    return f"{colour}[{bar}]{reset} {pct:5.1f}%"


def _print_round_dashboard(
    round_num: int,
    total_rounds: int,
    train_loss: float,
    train_acc: float,
    eval_loss: Optional[float],
    eval_acc: Optional[float],
    round_time: float,
    history: List[dict],
) -> None:
    """Print a compact per-round summary box with ASCII accuracy bars and a
    rolling history table of the last few rounds."""

    W = 65
    print("=" * W)
    # ── Round progress ──────────────────────────────────────────
    done_blocks = int(round_num / total_rounds * 20)
    todo_blocks = 20 - done_blocks
    round_bar = "█" * done_blocks + "░" * todo_blocks
    print(
        f"  Round [{round_bar}] {round_num:>3}/{total_rounds}   "
        f"  elapsed: {round_time:.1f}s"
    )
    print("-" * W)

    # ── Accuracy bars ────────────────────────────────────────────
    print(f"  Train  acc  {_acc_bar(train_acc)}")
    print(f"  Train  loss  {train_loss:.4f}")
    if eval_acc is not None and eval_loss is not None:
        print(f"  Test   acc  {_acc_bar(eval_acc)}")
        print(f"  Test   loss  {eval_loss:.4f}")

    # ── Delta vs previous round (if available) ───────────────────
    if len(history) >= 2:
        prev = history[-2]
        delta_train = (train_acc - prev["train_acc"]) * 100
        arrow_t = "▲" if delta_train >= 0 else "▼"
        col_t = "\033[92m" if delta_train >= 0 else "\033[91m"
        rst = "\033[0m"
        print(
            f"  Train  Δacc  {col_t}{arrow_t}{abs(delta_train):.2f}%{rst} vs previous round",
            end="",
        )
        if eval_acc is not None and prev.get("eval_acc") is not None:
            delta_eval = (eval_acc - prev["eval_acc"]) * 100
            arrow_e = "▲" if delta_eval >= 0 else "▼"
            col_e = "\033[92m" if delta_eval >= 0 else "\033[91m"
            print(
                f"   |   Test Δacc  {col_e}{arrow_e}{abs(delta_eval):.2f}%{rst} vs previous round",
                end="",
            )
        print()

    # ── Rolling history table (last 5 rounds) ───────────────────
    show = history[-5:] if len(history) > 0 else []
    if show:
        print("-" * W)
        has_eval = any(h.get("eval_acc") is not None for h in show)
        hdr = f"  {'Rnd':>4}  {'TrLoss':>8}  {'TrAcc':>7}"
        if has_eval:
            hdr += f"  {'TsLoss':>8}  {'TsAcc':>7}"
        hdr += f"  {'Time':>6}"
        print(hdr)
        print("  " + "-" * (len(hdr) - 2))
        for h in show:
            marker = " ◄" if h["round"] == round_num else ""
            row = f"  {h['round']:>4}  {h['train_loss']:>8.4f}  {h['train_acc'] * 100:>6.2f}%"
            if has_eval:
                tsl = (
                    f"{h['eval_loss']:>8.4f}"
                    if h.get("eval_loss") is not None
                    else f"{'—':>8}"
                )
                tsa = (
                    f"{h['eval_acc'] * 100:>6.2f}%"
                    if h.get("eval_acc") is not None
                    else f"{'—':>7}"
                )
                row += f"  {tsl}  {tsa}"
            row += f"  {h['time_s']:>5.1f}s{marker}"
            print(row)
    print("=" * W)


# ---------------------------------------------------------------------------
# FedProx aggregation (sample-weighted average)
# ---------------------------------------------------------------------------


def fedavg(
    client_weights: List[List[np.ndarray]],
    client_sizes: List[int],
) -> List[np.ndarray]:
    """
    Compute sample-weighted average of model parameters (FedAvg / FedProx share
    the same aggregation step on the server — the difference is in the LOCAL
    objective each client optimises).

    W_global = sum_i( n_i / N * W_i )
    """
    total = sum(client_sizes)
    if total == 0:
        raise ValueError("fedavg: total sample count is zero.")

    averaged: List[np.ndarray] = []
    for layer_idx in range(len(client_weights[0])):
        weighted = np.zeros_like(client_weights[0][layer_idx], dtype=np.float64)
        for i, weights in enumerate(client_weights):
            weighted += weights[layer_idx].astype(np.float64) * (
                client_sizes[i] / total
            )
        averaged.append(weighted.astype(client_weights[0][layer_idx].dtype))
    return averaged


# ---------------------------------------------------------------------------
# Checkpoint helpers
# ---------------------------------------------------------------------------


def load_checkpoint(model: MedicalCNN) -> bool:
    """Load models/global_model.pth into model in-place. Returns True if found."""
    latest = MODELS_DIR / "global_model.pth"
    if not latest.exists():
        return False
    model.load_state_dict(torch.load(latest, map_location=DEVICE))
    return True


def save_checkpoint(model: MedicalCNN, round_num: int) -> None:
    """Save three checkpoint files after each round."""
    ts_tag = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    state = model.state_dict()

    torch.save(state, MODELS_DIR / "global_model.pth")
    torch.save(state, MODELS_DIR / f"global_model_round_{round_num}.pth")
    torch.save(state, MODELS_DIR / f"global_model_round_{round_num}_{ts_tag}.pth")

    print(
        f"[{_ts()}] [Save] global_model.pth + round_{round_num} + round_{round_num}_{ts_tag}"
    )


# ---------------------------------------------------------------------------
# One FedProx round
# ---------------------------------------------------------------------------


def run_round(
    global_model: MedicalCNN,
    client_loaders: List[DataLoader],
    local_epochs: int,
    round_num: int,
    lr: float = DEFAULT_LR,
    mu: float = DEFAULT_MU,
) -> Tuple[MedicalCNN, float, float]:
    """
    Execute one complete FedProx round:
      1. Snapshot global weights (W_global) — used in proximal term.
      2. Give each client a copy of W_global.
      3. Each client trains locally using:
             L = CrossEntropyLoss + (mu/2)*||W_local - W_global||^2
         with SGD + CosineAnnealingLR.
      4. Aggregate with sample-weighted FedAvg.
      5. Write aggregated weights back into global_model.
    """
    global_params = get_parameters(global_model)  # frozen snapshot of W_global

    all_weights: List[List[np.ndarray]] = []
    all_sizes: List[int] = []
    all_losses: List[float] = []
    all_accs: List[float] = []

    for cid, loader in enumerate(client_loaders, start=1):
        t0 = time.perf_counter()

        # Fresh copy of the global model for this client
        local_model = MedicalCNN(num_classes=NUM_CLASSES)
        set_parameters(local_model, global_params)
        local_model = _try_compile(local_model)

        # Class-weighted loss — penalise rare classes more
        cw = compute_class_weights(loader)

        # FedProx local training
        loss, acc = train_fedprox(
            model=local_model,
            global_params=global_params,
            loader=loader,
            epochs=local_epochs,
            lr=lr,
            mu=mu,
            device=DEVICE,
            class_weights=cw,
            client_id=cid,
            round_num=round_num,
        )
        elapsed = time.perf_counter() - t0
        n = len(loader.dataset)

        print(
            f"[{_ts()}] [Client {cid}] "
            f"loss={loss:.4f}  acc={acc * 100:.2f}%  "
            f"n={n}  time={elapsed:.1f}s"
        )

        all_weights.append(get_parameters(local_model))
        all_sizes.append(n)
        all_losses.append(loss)
        all_accs.append(acc)

    # Server-side weighted aggregation (identical for FedAvg and FedProx)
    set_parameters(global_model, fedavg(all_weights, all_sizes))

    total = sum(all_sizes)
    round_loss = sum(l * s for l, s in zip(all_losses, all_sizes)) / total
    round_acc = sum(a * s for a, s in zip(all_accs, all_sizes)) / total
    print(
        f"[{_ts()}] [Server] aggregated loss={round_loss:.4f}  acc={round_acc * 100:.2f}%"
    )

    return global_model, round_loss, round_acc


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------


def run_evaluation(
    model: MedicalCNN,
    round_num: int,
) -> Tuple[float, float, Dict[int, float]]:
    """
    Evaluate model on the global test partition.
    Prints global loss, global accuracy, and per-class accuracy.
    Returns (loss, overall_acc, per_class_acc_dict).
    """
    nw = NUM_EVAL_WORKERS
    loader = get_test_loader(
        batch_size=128,
        num_workers=nw,
        pin_memory=(DEVICE == "cuda"),
        prefetch_factor=2 if nw > 0 else None,
        persistent_workers=nw > 0,
    )
    loss, acc, per_class = evaluate(
        model, loader, device=DEVICE, num_classes=NUM_CLASSES
    )
    n = len(loader.dataset)
    print(
        f"[{_ts()}] [Eval R{round_num:02d}] "
        f"loss={loss:.4f}  acc={acc * 100:.2f}%  ({n} images)"
    )
    print("  Per-class accuracy:")
    for i, cls in enumerate(CLASS_NAMES):
        print(f"    {cls:12s}: {per_class[i] * 100:6.2f}%")
    return loss, acc, per_class


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────


def export_excel(
    history: List[dict],
    args,
    init_acc: float,
    final_acc: float,
    final_loss: float,
    total_time: float,
    client_loaders,
    per_class_history: Optional[List[Optional[Dict[int, float]]]] = None,
) -> "Path | None":
    """
    Export full training results to an Excel workbook saved in models/.
    3 sheets:
      1. Parameters    -- all hyperparameters used for this run
      2. Round Metrics -- per-round loss / accuracy / time (best row highlighted)
      3. Summary       -- baseline vs final accuracy, delta, total wall-clock time
    """
    if not _EXCEL_AVAILABLE:
        print("[Excel] openpyxl not installed -- skipping Excel export.")
        print("        Install with: pip install openpyxl")
        return None

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = MODELS_DIR / f"training_report_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # Colour palette
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
    GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
    WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=13)
    BOLD_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")

    def _header_row(ws, row_idx: int, values: list):
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # ================================================================
    # Sheet 1: Parameters
    # ================================================================
    ws1 = wb.active
    ws1.title = "Parameters"

    ws1.merge_cells("A1:B1")
    t1 = ws1["A1"]
    t1.value = "Tecnomate FL -- Training Parameters"
    t1.font = TITLE_FONT
    t1.alignment = CENTER
    ws1.row_dimensions[1].height = 24

    _header_row(ws1, 2, ["Parameter", "Value"])

    params = [
        ("Run timestamp", ts),
        ("Algorithm", f"FedProx (mu={args.mu})"),
        ("Federated rounds", args.rounds),
        ("Local epochs per round", args.epochs),
        ("Batch size", args.batch),
        ("Number of FL clients", NUM_CLIENTS),
        ("Training device", DEVICE),
        ("Fresh start (no resume)", str(args.fresh)),
        ("Per-round eval enabled", str(args.eval)),
        ("Model architecture", "ResNet-18 (grayscale adapted, 192x192)"),
        ("ResNet-18 modifications", "conv1: 3ch->1ch  |  fc: 512->6"),
        ("Number of output classes", NUM_CLASSES),
        ("Class names", ", ".join(CLASS_NAMES)),
        ("Optimizer", f"SGD lr={args.lr} momentum=0.9 nesterov=True wd=1e-4"),
        ("LR scheduler", "CosineAnnealingLR T_max=epochs eta_min=lr*0.01"),
        ("Loss function", "CrossEntropyLoss (class-weighted) + FedProx proximal"),
        ("FL aggregation", "Sample-weighted FedAvg (server side)"),
        (
            "Image augmentation (train)",
            "CLAHE(p=0.5) + HFlip(0.5) + Rotate+-15 + Affine",
        ),
        ("WeightedRandomSampler", "Yes -- weight = 1/class_count per sample"),
        ("Gradient clipping", "max_norm=1.0"),
        ("Total training samples", sum(len(l.dataset) for l in client_loaders)),
    ]

    for i, (key, val) in enumerate(params, start=3):
        ws1.cell(row=i, column=1, value=key).font = BOLD_FONT
        ws1.cell(row=i, column=2, value=val)
        if i % 2 == 0:
            for col in (1, 2):
                ws1.cell(row=i, column=col).fill = ALT_FILL

    _auto_width(ws1)

    # ================================================================
    # Sheet 2: Round Metrics
    # ================================================================
    ws2 = wb.create_sheet("Round Metrics")

    has_pc = bool(
        args.eval and per_class_history and any(p for p in per_class_history if p)
    )
    pc_headers = (
        [f"{CLASS_NAMES[i]} Acc (%)" for i in range(NUM_CLASSES)] if has_pc else []
    )
    eval_headers = ["Test Loss", "Test Acc (%)"] if args.eval else []
    all_headers = (
        ["Round", "Train Loss", "Train Acc (%)"]
        + eval_headers
        + pc_headers
        + ["Time (s)", "Status"]
    )

    merge_end = get_column_letter(len(all_headers))
    ws2.merge_cells(f"A1:{merge_end}1")
    t2 = ws2["A1"]
    t2.value = "Tecnomate FL -- Per-Round Training Metrics"
    t2.font = TITLE_FONT
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 24
    _header_row(ws2, 2, all_headers)

    best_acc = max(h["train_acc"] for h in history)
    for i, h in enumerate(history, start=3):
        col = 1
        ws2.cell(row=i, column=col, value=h["round"])
        col += 1
        ws2.cell(row=i, column=col, value=round(h["train_loss"], 4))
        col += 1
        ws2.cell(row=i, column=col, value=round(h["train_acc"] * 100, 2))
        col += 1
        if args.eval:
            ws2.cell(
                row=i,
                column=col,
                value=round(h["eval_loss"], 4) if h["eval_loss"] is not None else "N/A",
            )
            col += 1
            ws2.cell(
                row=i,
                column=col,
                value=round(h["eval_acc"] * 100, 2)
                if h["eval_acc"] is not None
                else "N/A",
            )
            col += 1
            if has_pc and per_class_history and i - 3 < len(per_class_history):
                pc = per_class_history[i - 3] or {}
                for ci in range(NUM_CLASSES):
                    ws2.cell(row=i, column=col, value=round(pc.get(ci, 0.0) * 100, 2))
                    col += 1
        is_best = abs(h["train_acc"] - best_acc) < 1e-9
        ws2.cell(row=i, column=col, value=round(h["time_s"], 1))
        col += 1
        ws2.cell(row=i, column=col, value="* Best" if is_best else "")
        row_fill = GOOD_FILL if is_best else (ALT_FILL if i % 2 == 0 else None)
        for c in range(1, col + 1):
            cell = ws2.cell(row=i, column=c)
            cell.alignment = CENTER
            if row_fill:
                cell.fill = row_fill

    _auto_width(ws2)

    # ================================================================
    # Sheet 3: Summary
    # ================================================================
    ws3 = wb.create_sheet("Summary")

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value = "Tecnomate FL -- Training Summary"
    t3.font = TITLE_FONT
    t3.alignment = CENTER
    ws3.row_dimensions[1].height = 24

    _header_row(ws3, 2, ["Metric", "Value"])

    delta = (final_acc - init_acc) * 100
    arrow = "+" if delta >= 0 else ""
    summary_rows = [
        ("Baseline test accuracy", f"{init_acc * 100:.2f}%"),
        ("Final test accuracy", f"{final_acc * 100:.2f}%"),
        ("Final test loss", f"{final_loss:.4f}"),
        ("Improvement (delta)", f"{arrow}{delta:.2f}%"),
        ("Best round train acc", f"{best_acc * 100:.2f}%"),
        ("Total wall-clock time (s)", round(total_time, 1)),
        ("Total wall-clock time (min)", round(total_time / 60, 2)),
        ("Total FL rounds completed", args.rounds),
        ("Local epochs per round", args.epochs),
        ("Algorithm", f"FedProx (mu={args.mu})"),
        ("Optimizer", f"SGD lr={args.lr} momentum=0.9 nesterov"),
        ("Scheduler", "CosineAnnealingLR"),
        ("Model", "ResNet-18 grayscale 192x192"),
    ]
    for i, (key, val) in enumerate(summary_rows, start=3):
        ws3.cell(row=i, column=1, value=key).font = BOLD_FONT
        cell = ws3.cell(row=i, column=2, value=val)
        if "Improvement" in key:
            cell.fill = GOOD_FILL if delta >= 0 else WARN_FILL
        elif i % 2 == 0:
            for c in (1, 2):
                ws3.cell(row=i, column=c).fill = ALT_FILL

    _auto_width(ws3)

    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────


def main() -> None:
    global DEVICE
    parser = argparse.ArgumentParser(
        description="Tecnomate -- standalone FedProx simulation (no network required)",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--rounds", type=int, default=DEFAULT_ROUNDS, help="Federated rounds"
    )
    parser.add_argument(
        "--epochs",
        type=int,
        default=DEFAULT_EPOCHS,
        help="Local epochs per client per round",
    )
    parser.add_argument(
        "--batch", type=int, default=DEFAULT_BATCH, help="Mini-batch size"
    )
    parser.add_argument(
        "--lr", type=float, default=DEFAULT_LR, help="SGD learning rate"
    )
    parser.add_argument(
        "--mu",
        type=float,
        default=DEFAULT_MU,
        help="FedProx proximal coefficient (0=FedAvg)",
    )
    parser.add_argument(
        "--device",
        type=str,
        default=DEVICE,
        help="Training device: mps | cuda | cpu  (auto-detected by default)",
    )
    parser.add_argument(
        "--num-workers",
        type=int,
        default=DEFAULT_WORKERS,
        help="DataLoader worker processes per client",
    )
    parser.add_argument(
        "--fresh",
        action="store_true",
        help="Ignore checkpoint, start fresh",
    )
    parser.add_argument(
        "--eval",
        action="store_true",
        help="Evaluate on global test set every round",
    )
    args = parser.parse_args()

    # Apply device override so all helpers pick it up
    DEVICE = args.device

    # Startup banner
    _banner("Tecnomate  --  FedProx Simulation  (single-process)")
    print(f"  Algorithm     : FedProx (mu={args.mu})")
    print(f"  Rounds        : {args.rounds}")
    print(f"  Local epochs  : {args.epochs}")
    print(f"  Batch size    : {args.batch}")
    print(f"  Optimizer     : SGD lr={args.lr} momentum=0.9 nesterov=True")
    print(f"  Scheduler     : CosineAnnealingLR")
    print(f"  Model         : ResNet-18 (grayscale, 192x192)")
    print(f"  Clients       : {NUM_CLIENTS}")
    print(
        f"  Device        : {DEVICE}  {'(Apple MPS 🚀)' if DEVICE == 'mps' else '(CUDA 🚀)' if DEVICE == 'cuda' else '(CPU)'}"
    )
    nw_label = (
        args.num_workers if args.num_workers > 0 else "0 (in-process, fastest on MPS)"
    )
    print(f"  DataLoader    : {nw_label} workers per client")
    print(f"  Batch size    : {args.batch}  |  torch.compile: eager") 
    print(f"  Eval per round: {args.eval}")
    print(f"  Fresh start   : {args.fresh}")
    print("=" * 65)

    # Initialise global model
    global_model = MedicalCNN(num_classes=NUM_CLASSES)

    if args.fresh:
        print(f"\n[{_ts()}] [Init] Starting from random weights (--fresh)")
    else:
        if load_checkpoint(global_model):
            print(f"[{_ts()}] [Init] Resumed from models/global_model.pth")
        else:
            print(
                f"[{_ts()}] [Init] No checkpoint found -- starting from random weights"
            )

    # Load client data loaders
    print(f"\n[{_ts()}] [Data] Loading client partitions...")
    client_loaders: List[DataLoader] = []
    for cid in range(1, NUM_CLIENTS + 1):
        # pin_memory is unsupported on MPS; only enable for CUDA
        pin = DEVICE == "cuda"
        # prefetch_factor requires num_workers > 0
        nw = args.num_workers
        loader = get_client_loader(
            cid,
            batch_size=args.batch,
            num_workers=nw,
            pin_memory=pin,
            prefetch_factor=2 if nw > 0 else None,
            persistent_workers=nw > 0,
        )
        dist = loader.dataset.class_distribution()
        print(
            f"  Client {cid}: {len(loader.dataset):,} samples  |  "
            + "  ".join(f"{k}={v}" for k, v in dist.items())
        )
        client_loaders.append(loader)

    # Baseline evaluation
    print(f"\n[{_ts()}] [Eval R00] Baseline (before any training):")
    init_loss, init_acc, _ = run_evaluation(global_model, round_num=0)

    # Training loop
    history: List[dict] = []
    per_class_history: List[Optional[Dict[int, float]]] = []
    wall_start = time.perf_counter()

    for rnd in range(1, args.rounds + 1):
        _hline()
        print(f"[{_ts()}]  ROUND  {rnd} / {args.rounds}")
        _hline()

        t0 = time.perf_counter()
        global_model, r_loss, r_acc = run_round(
            global_model,
            client_loaders,
            args.epochs,
            rnd,
            lr=args.lr,
            mu=args.mu,
        )
        round_time = time.perf_counter() - t0

        # Save checkpoint after every round -- partial progress is never lost
        save_checkpoint(global_model, rnd)

        # Optional per-round test evaluation
        eval_loss, eval_acc, eval_per_class = None, None, None
        if args.eval:
            eval_loss, eval_acc, eval_per_class = run_evaluation(
                global_model, round_num=rnd
            )

        history.append(
            {
                "round": rnd,
                "train_loss": r_loss,
                "train_acc": r_acc,
                "eval_loss": eval_loss,
                "eval_acc": eval_acc,
                "time_s": round_time,
            }
        )
        per_class_history.append(eval_per_class)

        # ── Per-round dashboard ───────────────────────────────────
        _print_round_dashboard(
            round_num=rnd,
            total_rounds=args.rounds,
            train_loss=r_loss,
            train_acc=r_acc,
            eval_loss=eval_loss,
            eval_acc=eval_acc,
            round_time=round_time,
            history=history,
        )

    total_time = time.perf_counter() - wall_start

    # Final evaluation
    _banner("Training Complete  --  Final Evaluation")
    final_loss, final_acc, final_per_class = run_evaluation(
        global_model, round_num=args.rounds
    )

    # Summary table
    print()
    _banner("Training Summary")

    header = (
        f"  {'Round':>6}  {'Train Loss':>10}  {'Train Acc':>10}"
        + (f"  {'Test Loss':>10}  {'Test Acc':>10}" if args.eval else "")
        + f"  {'Time(s)':>8}"
    )
    print(header)
    print("  " + "-" * (len(header) - 2))

    for h in history:
        row = f"  {h['round']:>6}  {h['train_loss']:>10.4f}  {h['train_acc'] * 100:>9.2f}%"
        if args.eval and h["eval_acc"] is not None:
            row += f"  {h['eval_loss']:>10.4f}  {h['eval_acc'] * 100:>9.2f}%"
        row += f"  {h['time_s']:>8.1f}"
        print(row)

    print()
    print(f"  Baseline test accuracy  : {init_acc * 100:.2f}%")
    print(f"  Final   test accuracy   : {final_acc * 100:.2f}%")
    delta = (final_acc - init_acc) * 100
    arrow = "+" if delta >= 0 else ""
    print(f"  Delta                   : {arrow}{delta:.2f}%")
    print(f"  Total wall-clock time   : {total_time:.1f}s  ({total_time / 60:.1f} min)")
    print()
    print(f"  Best model saved to     : models/global_model.pth")
    print(f"  Run for full report     : python src/evaluate_global.py")
    print("=" * 65)

    # Excel export
    excel_path = export_excel(
        history=history,
        args=args,
        init_acc=init_acc,
        final_acc=final_acc,
        final_loss=final_loss,
        total_time=total_time,
        client_loaders=client_loaders,
        per_class_history=per_class_history,
    )
    if excel_path:
        print(f"  Excel report saved to   : models/{excel_path.name}")
        print("=" * 65)


if __name__ == "__main__":
    main()
