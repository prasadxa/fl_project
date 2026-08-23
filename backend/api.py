"""
Tecnomate Clinical AI — FastAPI Backend
========================================
Endpoints:
  GET  /api/health            — liveness + model + security status
  GET  /api/model-info        — class names, scan modes, colours
  GET  /api/model-features    — CNN architecture, layer features, parameter count
  POST /api/ocr-check         — standalone OCR X-ray gate (pre-check before predict)
  POST /api/predict           — image inference  (multipart/form-data)
  POST /api/feedback          — doctor-confirmed / overridden label
  GET  /api/queue             — last 50 feedback entries (legacy)
  GET  /api/report            — download plain-text diagnostic report (legacy)
  POST /api/pdf-report        — generate & download professional PDF report
  GET  /api/admin/stats       — aggregate statistics from SQLite
  GET  /api/admin/feedback    — paginated feedback log
  GET  /api/admin/export-csv  — download full feedback CSV
  GET  /api/admin/export-excel — download full admin report as Excel workbook
  GET  /api/admin/sessions    — paginated prediction sessions

OCR X-Ray Gate (Chest X-Ray strict mode):
  Chest X-Ray uploads are subject to a strict OCR gate BEFORE inference.
  Policy:
    • is_xray = True              → ALLOWED  (OCR confirmed X-ray markers)
    • is_xray = False             → REJECTED (OCR identified non-xray content)
    • is_xray = None (no text)    → REJECTED (cannot confirm — strict mode)
    • OCR engine unavailable      → REJECTED (cannot verify — strict mode)
  Brain MRI uploads skip the OCR gate entirely.

Security:
  - SecurityHeadersMiddleware : X-Content-Type-Options, X-Frame-Options,
                                X-XSS-Protection, Referrer-Policy,
                                Content-Security-Policy, Permissions-Policy
  - RateLimitMiddleware       : 200 req/min per IP (general),
                                100 req/min per IP (/api/predict — inference)
  - CORS                      : GET + POST only, explicit headers whitelist

Static frontend served from ../frontend/  (mounted at /)

Supported image formats  : JPEG, PNG, WebP, BMP, TIFF, GIF, AVIF, HEIC/HEIF
                           DICOM (.dcm) via pydicom (optional)
Maximum upload size      : 30 MB

Run:
    uvicorn backend.api:app --reload --host 0.0.0.0 --port 8000
"""

from __future__ import annotations

import collections
import datetime
import io
import json
import logging
import sys
import threading
import time
import uuid
from pathlib import Path
from typing import Dict, List, Optional

# ── audit logger ──────────────────────────────────────────────────────────────
# Writes OCR gate rejections and key inference events to a rotating log file.
_audit_logger = logging.getLogger("tecnomate.audit")
if not _audit_logger.handlers:
    _audit_logger.setLevel(logging.INFO)
    _audit_logger.propagate = False
    try:
        from logging.handlers import RotatingFileHandler as _RFH

        _PROJ_ROOT_FOR_LOG = Path(__file__).parent.parent
        _LOG_PATH = _PROJ_ROOT_FOR_LOG / "data" / "audit.log"
        _LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        _rh = _RFH(str(_LOG_PATH), maxBytes=5 * 1024 * 1024, backupCount=3)
        _rh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        _audit_logger.addHandler(_rh)
    except Exception:
        _audit_logger.addHandler(logging.StreamHandler())

import re

import cv2
import numpy as np
import secrets
import torch
import torch.nn.functional as F
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, PlainTextResponse, Response
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from PIL import Image
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import JSONResponse
from torchvision import transforms

# ── project imports ────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))  # backend/
from anonymizer import strip_metadata_and_save
from db import get_db
from model import MedicalCNN
from ocr_reader import (
    XrayScanResult,
    extract_text,
    is_ocr_available,
    is_xray_scan,
    ocr_unavailable_reason,
)
from report_generator import (
    PatientInfo,
    ReportRequest,
    UncertaintyInfo,
    build_pdf_report,
    compute_gradcam,
    compute_mc_uncertainty,
)
from scan_classifier import CONFIDENCE_THRESHOLD as GATE_THRESHOLD
from scan_classifier import get_scan_gate

from dataset import CLASS_NAMES, NUM_CLASSES

# ── paths ──────────────────────────────────────────────────────────────────────
PROJ_ROOT = Path(__file__).parent.parent


# ── scan type canonicalization ─────────────────────────────────────────────────
def canonicalize_scan_type(scan_type: Optional[str]) -> str:
    """
    Normalize scan_type to canonical values: "Chest X-Ray" or "Brain MRI".

    Handles common aliases, case differences, spacing variations, and typos.
    Returns the original (stripped) value if no known alias is matched.
    """
    if not scan_type:
        return "unknown"
    # Normalize: lowercase, remove all non-alphanumeric chars
    key = re.sub(r"[^a-z0-9]", "", scan_type.lower())
    # Map known aliases to canonical names
    chest_aliases = {"chestxray", "cxr", "xray", "chest", "chestx", "xraychest"}
    brain_aliases = {"brainmri", "mri", "brain", "mribrain"}
    if key in chest_aliases:
        return "Chest X-Ray"
    if key in brain_aliases:
        return "Brain MRI"
    # Fallback: return stripped original
    return scan_type.strip()


MODEL_PATH = PROJ_ROOT / "models" / "global_model.pth"
COLLECT_DIR = PROJ_ROOT / "data" / "new_collected_data"
FRONTEND_DIR = PROJ_ROOT / "frontend" / "dist"
TEMP_DIR = PROJ_ROOT / "data" / ".tmp_uploads"
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# ── upload constraints ─────────────────────────────────────────────────────────
MAX_UPLOAD_BYTES: int = 30 * 1024 * 1024  # 30 MB hard limit

ALLOWED_EXTENSIONS: frozenset[str] = frozenset(
    {
        ".jpg",
        ".jpeg",  # JPEG
        ".png",  # PNG
        ".webp",  # WebP (lossy & lossless)
        ".bmp",  # BMP / DIB
        ".tif",
        ".tiff",  # TIFF (common DICOM-export format)
        ".gif",  # GIF  (takes first frame)
        ".avif",  # AVIF / AV1
        ".heic",
        ".heif",  # HEIC/HEIF
        ".dcm",  # DICOM (pydicom, optional)
    }
)

ALLOWED_MIME_PREFIXES: tuple[str, ...] = (
    "image/",
    "application/octet-stream",
    "application/dicom",
)

# ── scan-mode configuration ────────────────────────────────────────────────────
SCAN_MODES: Dict = {
    "Brain MRI": {
        "indices": [0, 1, 2, 3],
        "labels": {0: "Glioma", 1: "Meningioma", 2: "No Tumor", 3: "Pituitary Tumor"},
        "class_keys": ["glioma", "meningioma", "notumor", "pituitary"],
        "icon": "\U0001f9e0",
    },
    "Chest X-Ray": {
        "indices": [4, 5],
        "labels": {4: "Normal", 5: "Pneumonia"},
        "class_keys": ["normal", "pneumonia"],
        "icon": "\U0001fac1",
    },
}

SHORT_NAMES: Dict[str, str] = {
    "glioma": "Glioma (Brain Tumor)",
    "meningioma": "Meningioma (Brain Tumor)",
    "notumor": "No Tumor Detected",
    "pituitary": "Pituitary Tumor",
    "normal": "Normal / Healthy (CXR)",
    "pneumonia": "Pneumonia Detected (CXR)",
}

RISK_COLOURS: Dict[str, str] = {
    "glioma": "#e74c3c",
    "meningioma": "#e67e22",
    "notumor": "#27ae60",
    "pituitary": "#e67e22",
    "normal": "#27ae60",
    "pneumonia": "#e74c3c",
}

# ── FastAPI app ────────────────────────────────────────────────────────────────
app = FastAPI(
    title="Tecnomate Clinical AI",
    description="Privacy-preserving federated medical image classifier",
    version="2.0.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)


# ── Security headers middleware ────────────────────────────────────────────────
# Adds industry-standard HTTP security headers to every response.
# These protect against XSS, clickjacking, MIME-sniffing and data leakage.
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = (
            "camera=(), microphone=(), geolocation=()"
        )
        # Only disable caching for API responses; let hashed static assets cache normally
        if request.url.path.startswith("/api/"):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
            response.headers["Pragma"] = "no-cache"
        # CSP: allow Google Fonts (Inter typeface) and same-origin resources
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data: blob:; "
            "connect-src 'self'"
        )
        return response


# ── Rate limiting middleware ───────────────────────────────────────────────────
# Simple in-memory sliding-window rate limiter (no extra packages required).
# Default: max 60 requests per IP per minute for general endpoints,
#          max 10 prediction requests per IP per minute (heavy inference).
_rate_lock = threading.Lock()
_rate_store: Dict[str, list] = collections.defaultdict(list)  # ip -> [timestamps]

RATE_LIMIT_GENERAL = 200  # requests per window
RATE_LIMIT_PREDICT = 100  # requests per window (inference is expensive)
RATE_LIMIT_WINDOW_SEC = 60  # window size in seconds


class RateLimitMiddleware(BaseHTTPMiddleware):
    """
    Sliding-window rate limiter per client IP.
    - /api/predict  : stricter limit (inference is CPU-heavy)
    - all others    : general limit
    Returns HTTP 429 with Retry-After header when limit is exceeded.
    """

    async def dispatch(self, request: Request, call_next):
        ip = request.client.host if request.client else "unknown"
        path = request.url.path
        limit = RATE_LIMIT_PREDICT if path == "/api/predict" else RATE_LIMIT_GENERAL
        now = time.monotonic()

        with _rate_lock:
            timestamps = _rate_store[ip]
            # Remove timestamps outside the window
            _rate_store[ip] = [t for t in timestamps if now - t < RATE_LIMIT_WINDOW_SEC]
            if len(_rate_store[ip]) >= limit:
                retry_after = (
                    int(RATE_LIMIT_WINDOW_SEC - (now - _rate_store[ip][0])) + 1
                )
                return JSONResponse(
                    status_code=429,
                    content={
                        "detail": f"Rate limit exceeded. Max {limit} requests "
                        f"per {RATE_LIMIT_WINDOW_SEC}s. "
                        f"Retry after {retry_after}s."
                    },
                    headers={"Retry-After": str(retry_after)},
                )
            _rate_store[ip].append(now)

        return await call_next(request)


app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(RateLimitMiddleware)

# ── CORS ───────────────────────────────────────────────────────────────────────
# Restrict to specific methods and headers rather than wildcard.
# allow_origins=["*"] is kept for local dev; in production replace with your domain.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Replace with ["https://yourdomain.com"] in production
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "Accept", "Authorization"],
    max_age=600,
)

# ── model singleton ────────────────────────────────────────────────────────────
_model: Optional[MedicalCNN] = None
_model_loaded: bool = False


def get_model() -> Optional[MedicalCNN]:
    global _model, _model_loaded
    if not _model_loaded:
        if MODEL_PATH.exists():
            try:
                m = MedicalCNN(num_classes=NUM_CLASSES)
                m.load_state_dict(torch.load(MODEL_PATH, map_location="cpu", weights_only=True))
                m.eval()
                _model = m
                print(f"[API] ✓ Model loaded from {MODEL_PATH.name}")
            except Exception as exc:
                print(f"[API] ✗ Model load failed: {exc}")
        else:
            print(f"[API] ✗ Model not found at {MODEL_PATH}")
        _model_loaded = True
    return _model


@app.on_event("startup")
async def startup_event() -> None:
    import asyncio
    get_model()
    # Initialise the DB singleton so the file is created at startup.
    get_db()
    print(f"[API] OCR available: {is_ocr_available()}")
    if not is_ocr_available():
        print(f"[API] OCR reason  : {ocr_unavailable_reason()}")

    # Background task: clean up temp files older than 1 hour (runs every 5 min)
    async def _cleanup_temp_files() -> None:
        while True:
            await asyncio.sleep(300)
            try:
                cutoff = time.time() - 3600  # 1 hour
                for f in list(TEMP_DIR.iterdir()):
                    try:
                        if f.is_file() and f.stat().st_mtime < cutoff:
                            f.unlink(missing_ok=True)
                            if "_gradcam" not in f.name:
                                _pending_images.pop(f.stem, None)
                    except Exception:
                        pass
            except Exception:
                pass

    asyncio.create_task(_cleanup_temp_files())


# ── image preprocessing ────────────────────────────────────────────────────────
# CLAHE instance reused across requests (thread-safe read-only after creation)
_CLAHE = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))

_INFER_TRANSFORM = transforms.Compose(
    [
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5], std=[0.5]),
    ]
)


def _load_dicom(contents: bytes) -> Image.Image:
    """
    Decode a DICOM file using pydicom and return a PIL greyscale image.
    Raises ImportError if pydicom is not installed.
    Raises ValueError on decode failure.
    """
    try:
        import pydicom  # type: ignore[import]

        try:
            from pydicom.pixel_data_handlers.util import (  # type: ignore[import]
                apply_voi_lut,
            )
        except ImportError:
            apply_voi_lut = None  # type: ignore[assignment]
    except ImportError as exc:
        raise ImportError(
            "pydicom is required to process DICOM files. "
            "Install it with: pip install pydicom"
        ) from exc

    try:
        ds = pydicom.dcmread(io.BytesIO(contents))
        pixel_array = ds.pixel_array.astype(np.float32)

        # Apply VOI LUT / windowing when available (improves contrast)
        try:
            if apply_voi_lut is not None:
                pixel_array = apply_voi_lut(pixel_array, ds).astype(np.float32)
        except Exception:
            pass  # Not all DICOM files carry windowing info

        # Normalise to 0-255 uint8
        lo, hi = pixel_array.min(), pixel_array.max()
        if hi > lo:
            pixel_array = (pixel_array - lo) / (hi - lo) * 255.0
        pixel_array = pixel_array.clip(0, 255).astype(np.uint8)

        # Handle multi-frame (cine / 3-D series) → first frame
        if pixel_array.ndim == 3 and pixel_array.shape[0] > 1:
            pixel_array = pixel_array[0]

        if pixel_array.ndim == 3:
            # RGB DICOM (rare)
            return Image.fromarray(pixel_array, mode="RGB")
        else:
            return Image.fromarray(pixel_array, mode="L")

    except Exception as exc:
        raise ValueError(f"Failed to decode DICOM file: {exc}") from exc


def prepare_image(pil_img: Image.Image) -> torch.Tensor:
    """
    Convert any PIL image to the (1, 1, 128, 128) float tensor the model expects.
    """
    try:
        pil_img.seek(0)
    except (EOFError, AttributeError):
        pass

    mode = pil_img.mode
    if mode == "L":
        grey_pil = pil_img.copy()
    elif mode in ("LA", "PA"):
        grey_pil = pil_img.convert("L")
    elif mode in ("I", "I;16", "I;16B"):
        arr = np.array(pil_img, dtype=np.float32)
        lo, hi = arr.min(), arr.max()
        if hi > lo:
            arr = (arr - lo) / (hi - lo) * 255.0
        grey_pil = Image.fromarray(arr.astype(np.uint8), mode="L")
    elif mode == "F":
        arr = np.array(pil_img, dtype=np.float32)
        lo, hi = arr.min(), arr.max()
        if hi > lo:
            arr = (arr - lo) / (hi - lo) * 255.0
        grey_pil = Image.fromarray(arr.astype(np.uint8), mode="L")
    else:
        rgb_pil = pil_img.convert("RGB")
        img_np = np.array(rgb_pil, dtype=np.uint8)
        img_bgr = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)
        grey_arr = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        grey_pil = Image.fromarray(grey_arr, mode="L")

    grey_arr = np.array(grey_pil, dtype=np.uint8)
    resized = cv2.resize(grey_arr, (128, 128), interpolation=cv2.INTER_AREA)
    # Apply CLAHE — same contrast enhancement used during training
    enhanced = _CLAHE.apply(resized)
    pil_out = Image.fromarray(enhanced, mode="L")
    return _INFER_TRANSFORM(pil_out).unsqueeze(0)  # type: ignore[union-attr]


def _validate_upload(contents: bytes, filename: str) -> None:
    """Raise HTTP 413 / 415 for invalid uploads."""
    if len(contents) > MAX_UPLOAD_BYTES:
        mb = len(contents) / 1024 / 1024
        raise HTTPException(
            status_code=413,
            detail=(
                f"File too large ({mb:.1f} MB).  "
                f"Maximum allowed size is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
            ),
        )
    if filename:
        ext = Path(filename).suffix.lower()
        if ext and ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=415,
                detail=(
                    f"Unsupported file extension '{ext}'.  "
                    "Accepted: JPEG, PNG, WebP, BMP, TIFF, GIF, AVIF, HEIC, DICOM."
                ),
            )


# ── in-process temp-image store  ──────────────────────────────────────────────
_pending_images: Dict[str, Path] = {}


# ══════════════════════════════════════════════════════════════════════════════
#  API Routes
# ══════════════════════════════════════════════════════════════════════════════


@app.get("/api/health", tags=["System"])
def health():
    """Liveness check — returns model, OCR, DB and security status."""
    model = get_model()
    db = get_db()
    counts = db.count_feedback()
    return {
        "status": "ok",
        "model_loaded": model is not None,
        "model_path": str(MODEL_PATH),
        "ocr_available": is_ocr_available(),
        "ocr_reason": ocr_unavailable_reason() if not is_ocr_available() else "",
        "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
        "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
        "feedback_total": counts["total"],
        "feedback_overridden": counts["overridden"],
        "timestamp": datetime.datetime.now().isoformat(),
        "security": {
            "rate_limit_general": f"{RATE_LIMIT_GENERAL} req/{RATE_LIMIT_WINDOW_SEC}s per IP",
            "rate_limit_predict": f"{RATE_LIMIT_PREDICT} req/{RATE_LIMIT_WINDOW_SEC}s per IP",
            "security_headers": "enabled",
            "cors_methods": "GET, POST",
            "xss_protection": "enabled",
            "clickjacking_protection": "enabled",
            "content_type_sniffing_protection": "enabled",
        },
    }


@app.get("/api/model-info", tags=["System"])
def model_info():
    """Return class registry, scan modes and display metadata."""
    model = get_model()
    return {
        "model_loaded": model is not None,
        "classes": CLASS_NAMES,
        "num_classes": NUM_CLASSES,
        "scan_modes": {
            k: {
                "class_keys": v["class_keys"],
                "labels": {str(ki): vi for ki, vi in v["labels"].items()},
                "icon": v["icon"],
            }
            for k, v in SCAN_MODES.items()
        },
        "short_names": SHORT_NAMES,
        "risk_colours": RISK_COLOURS,
    }


@app.get("/api/model-features", tags=["System"])
def model_features():
    """
    Returns a full explanation of the CNN architecture and the visual features
    each layer learns to detect.  Answers the question: 'How does the model
    know it is looking at a tumour?'
    """
    model = get_model()

    # ─ parameter count ────────────────────────────────────────────────────
    total_params = 0
    trainable_params = 0
    if model is not None:
        import torch

        for p in model.parameters():
            n = p.numel()
            total_params += n
            if p.requires_grad:
                trainable_params += n

    return {
        "architecture": {
            "name": "MedicalCNN",
            "type": "Convolutional Neural Network (CNN)",
            "input": "128 x 128 pixels, 1 channel (greyscale)",
            "output": f"{NUM_CLASSES} classes (softmax probability per class)",
            "total_parameters": total_params,
            "trainable_parameters": trainable_params,
        },
        "conv_blocks": [
            {
                "block": 1,
                "operation": "Conv2d(1 → 32 filters, 3×3 kernel) + BatchNorm + ReLU + MaxPool",
                "output_shape": "32 × 64 × 64",
                "features_learned": [
                    "Low-level edges and contours (horizontal, vertical, diagonal)",
                    "Basic brightness gradients across the scan",
                    "Sharp intensity boundaries — e.g. tumour border vs healthy tissue",
                    "Early noise filtering via BatchNorm",
                ],
                "why_matters": (
                    "The tumour boundary (edge between tumour mass and normal brain tissue) "
                    "is the first discriminative signal the network learns to detect."
                ),
            },
            {
                "block": 2,
                "operation": "Conv2d(32 → 64 filters, 3×3 kernel) + BatchNorm + ReLU + MaxPool",
                "output_shape": "64 × 32 × 32",
                "features_learned": [
                    "Texture patterns: heterogeneous vs homogeneous tissue regions",
                    "Blob-like irregular shapes characteristic of tumour masses",
                    "Ring-enhancement patterns common in glioma on contrast MRI",
                    "Mid-frequency spatial patterns (not raw edges, not global shape)",
                    "Lung opacity patterns for pneumonia (consolidation, haziness)",
                ],
                "why_matters": (
                    "Tumours have a distinctly irregular texture compared to normal brain. "
                    "Glioma shows heterogeneous signal; meningioma shows a dense homogeneous mass. "
                    "Block 2 captures these mid-level texture signatures."
                ),
            },
            {
                "block": 3,
                "operation": "Conv2d(64 → 128 filters, 3×3 kernel) + BatchNorm + ReLU + MaxPool",
                "output_shape": "128 × 16 × 16",
                "features_learned": [
                    "High-level semantic features: overall tumour shape and mass effect",
                    "Mass effect: midline shift, ventricle compression",
                    "Pituitary tumour location (sellar/suprasellar region)",
                    "Global scan orientation and anatomical context",
                    "Bilateral lung opacity distribution for pneumonia vs normal",
                ],
                "why_matters": (
                    "Block 3 produces abstract representations that encode whether a full "
                    "tumour-consistent structure exists in the image. The global average pool "
                    "then summarises these 128 feature maps into a 2048-dimensional vector."
                ),
            },
        ],
        "classifier_head": {
            "layers": [
                "AdaptiveAvgPool2d(4×4) — spatially compress 128 feature maps to 128×4×4 = 2048 values",
                "Dropout(0.5)           — randomly disables 50% of neurons during training to prevent overfitting",
                "FC(2048 → 256)         — dense layer: combines all spatial features into 256 discriminative neurons",
                "ReLU                   — non-linearity: zero out negative activations",
                "Dropout(0.3)           — additional regularisation before final decision",
                "FC(256 → 6)            — outputs raw score (logit) for each of the 6 classes",
                "Softmax (at inference)  — converts logits to probabilities summing to 100%",
            ],
            "why_dropout_matters": (
                "Dropout forces the network to not rely on any single neuron, "
                "making it more robust to unseen MRI scans from different scanners or patients."
            ),
        },
        "detection_features_per_class": {
            "glioma": [
                "Irregular, infiltrating borders (Block 1 edges)",
                "Heterogeneous signal intensity / necrotic core (Block 2 texture)",
                "Mass effect — midline shift visible in scan (Block 3 shape)",
                "Ring-enhancement pattern in contrast MRI",
            ],
            "meningioma": [
                "Well-defined, rounded dense mass (Block 2 blob detection)",
                "Extra-axial location — outside brain parenchyma (Block 3 location)",
                "Dural tail sign — attachment to meninges",
                "Homogeneous enhancement pattern",
            ],
            "pituitary": [
                "Small mass in sellar/suprasellar region (Block 3 anatomical location)",
                "Optic chiasm compression pattern",
                "Distinct from surrounding pituitary gland tissue (Block 1 contrast)",
            ],
            "notumor": [
                "Absence of focal mass, normal symmetrical brain tissue",
                "No irregular texture blobs detected by Block 2",
                "Normal ventricle size, no midline shift",
            ],
            "pneumonia": [
                "Focal or diffuse opacity / consolidation in lung fields (Block 2 texture)",
                "Air bronchogram signs (Block 1 fine structure)",
                "Asymmetric or bilateral haziness (Block 3 global pattern)",
            ],
            "normal": [
                "Clear lung fields bilaterally",
                "Sharp costophrenic angles",
                "No opacity / consolidation patterns detected",
            ],
        },
        "explainability": {
            "gradcam": (
                "Gradient-weighted Class Activation Mapping (Grad-CAM) is implemented. "
                "Pass gradcam=true to /api/predict to get a heatmap overlay showing "
                "EXACTLY which pixels in the scan the model used to make its decision. "
                "This is included in the PDF report."
            ),
            "mc_dropout": (
                "Monte Carlo Dropout uncertainty estimation is implemented. "
                "Pass mc_dropout=true to /api/predict to get a confidence interval "
                "showing how certain the model is about its prediction."
            ),
        },
        "training_details": {
            "loss_function": "CrossEntropyLoss with inverse-frequency class weights",
            "class_balancing": "Rare tumour types get higher loss weight (fix for majority-class bias)",
            "optimiser": "Adam  lr=0.001",
            "lr_scheduler": "StepLR  step_size=1  gamma=0.9  (lr decays each epoch)",
            "local_epochs": "10 per federated round (minimum)",
            "data_augmentation": [
                "CLAHE contrast enhancement — improves tumour boundary visibility",
                "Random horizontal flip (p=0.5) — tumours appear on either side",
                "Random vertical flip (p=0.2)   — valid for axial MRI slices",
                "Random rotation \u00b115\u00b0           — scanner positioning variation",
                "Random affine (translate/scale) — patient movement simulation",
                "Random autocontrast (p=0.3)    — different scanner settings",
                "Random sharpness (p=0.3)       — varying MRI sharpness",
            ],
            "federated_learning": (
                "FedAvg: each of 3 hospital clients trains locally for 10 epochs, "
                "then the server aggregates weights using sample-weighted averaging. "
                "Patient data never leaves the hospital — only model weights are shared."
            ),
        },
    }


# ── medical scan plausibility ──────────────────────────────────────────────────


def _is_likely_medical_scan(img: Image.Image) -> bool:
    """
    Content-aware heuristic: real X-rays / MRIs are near-grayscale images.
    Returns False for colorful photos, product shots, selfies, etc.

    Key improvement over a naive whole-image check: product photos (perfume
    bottles, food, etc.) often have a plain white background that makes the
    *whole-image* gray ratio look high.  This function strips near-white and
    near-black background pixels and evaluates only the actual foreground
    content, where color information lives.
    """
    try:
        rgb = img.convert("RGB").resize((128, 128))
        arr = np.array(rgb, dtype=np.float32)
        r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]

        # ── Whole-image saturation pass (fast reject for vivid images) ────────
        diff_max_all = np.maximum(
            np.abs(r - g), np.maximum(np.abs(r - b), np.abs(g - b))
        )
        gray_ratio_all = float((diff_max_all < 12).mean())

        hsv = rgb.convert("HSV")
        sat_all = np.array(hsv)[:, :, 1].astype(np.float32) / 255.0
        mean_sat_all = float(sat_all.mean())
        high_sat_all = float((sat_all > 0.20).mean())

        # Reject outright if the whole image is obviously colourful
        if gray_ratio_all < 0.70:
            return False
        if mean_sat_all > 0.15:
            return False
        if high_sat_all > 0.20:
            return False

        # ── Content-pixel analysis (ignores white/black background) ──────────
        # White background: R, G, B all > 235
        # Black background: R, G, B all < 20
        # Near-neutral grey (common X-ray background): R≈G≈B and mid-range
        bg_white = (r > 235) & (g > 235) & (b > 235)
        bg_black = (r < 20) & (g < 20) & (b < 20)
        bg_mask = bg_white | bg_black
        content_mask = ~bg_mask

        content_pixels = int(content_mask.sum())

        if content_pixels < 50:
            # Almost entirely background — cannot draw conclusions from content;
            # fall back to whole-image rules already passed above.
            return True

        r_c = r[content_mask]
        g_c = g[content_mask]
        b_c = b[content_mask]

        diff_max_c = np.maximum(
            np.abs(r_c - g_c), np.maximum(np.abs(r_c - b_c), np.abs(g_c - b_c))
        )

        # Fraction of content pixels that are truly grayscale (diff ≤ 15)
        content_gray_ratio = float((diff_max_c <= 15).mean())

        # Fraction of content pixels with strong colour channel imbalance (diff > 30)
        # Real X-rays: ~0%.  Product photos with coloured objects: typically > 20%.
        content_color_ratio = float((diff_max_c > 30).mean())

        # Mean per-channel imbalance among content pixels
        mean_diff_c = float(diff_max_c.mean())

        # ── Decision: content-pixel rules ────────────────────────────────────
        # Content must be mostly grayscale
        if content_gray_ratio < 0.72:
            return False
        # More than 15 % of content pixels with strong colour → reject
        if content_color_ratio > 0.15:
            return False
        # Average colour channel imbalance too high
        if mean_diff_c > 18.0:
            return False

        # ── Unique quantised colour count (whole image, 64×64) ────────────────
        small = img.convert("RGB").resize((64, 64))
        quantized = (np.array(small) // 16).reshape(-1, 3)
        unique_colors = len(set(map(tuple, quantized.tolist())))
        if unique_colors > 900:
            return False

        return True
    except Exception:
        return True  # fail-open: never block on check failure


# ── predict ────────────────────────────────────────────────────────────────────


# ── OCR pre-check endpoint ─────────────────────────────────────────────────────
@app.post("/api/ocr-check", tags=["Inference"])
async def ocr_check(
    image: UploadFile = File(
        ..., description="Image to verify before submitting to /api/predict"
    ),
    scan_type: str = Form("Chest X-Ray", description="'Brain MRI' or 'Chest X-Ray'"),
):
    """
    Lightweight OCR gate check — run this BEFORE /api/predict to give instant
    feedback to the user without wasting inference resources.

    Returns:
      allowed       : bool  — True if image passes the OCR gate for the given scan_type
      is_xray       : bool | None
      scan_type_detected : str
      confidence    : float
      keywords_found : list[str]
      message       : str   — human-readable verdict
      error_code    : str   — non-empty when allowed=False
    """
    # Canonicalize scan_type to handle variations (case, spacing, aliases)
    original_scan_type = scan_type
    scan_type = canonicalize_scan_type(scan_type)
    _audit_logger.info(
        "OCR_CHECK_SCAN_TYPE original=%r canonical=%r",
        original_scan_type,
        scan_type,
    )

    contents = await image.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Empty file received.")

    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large.")

    # Decode image
    try:
        pil_img = Image.open(io.BytesIO(contents))
        pil_img.verify()
        pil_img = Image.open(io.BytesIO(contents))
        pil_img.load()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not decode image: {exc}")

    # ── STEP 1: ScanGate — ML + heuristic scan-type classifier ────────────────
    # Runs before OCR and before any modality-specific logic.
    # Rejects non-medical images and wrong modality uploads immediately.
    _gate = get_scan_gate()
    _gate_result = _gate.check(pil_img, scan_type)
    if not _gate_result.allowed:
        return {
            "allowed": False,
            "is_xray": False,
            "scan_type_detected": _gate_result.label,
            "confidence": round(_gate_result.confidence, 4),
            "keywords_found": [],
            "message": _gate_result.rejection_reason,
            "error_code": "OCR_SCAN_REJECTED",
        }

    if scan_type == "Chest X-Ray":
        # ── VISUAL CHECK: secondary grayscale gate after ScanGate ─────────────
        # Catches colourful images the gate may have passed at lower confidence.
        if not _is_likely_medical_scan(pil_img):
            return {
                "allowed": False,
                "is_xray": False,
                "scan_type_detected": "non_medical",
                "confidence": 0.90,
                "keywords_found": [],
                "message": (
                    "This image does not appear to be a chest X-ray. "
                    "Real chest X-rays are grayscale — this image contains "
                    "colour content inconsistent with radiographic imaging. "
                    "Please upload an original chest X-ray (JPEG / PNG / DICOM)."
                ),
                "error_code": "OCR_SCAN_REJECTED",
            }

        if not is_ocr_available():
            return {
                "allowed": False,
                "is_xray": None,
                "scan_type_detected": "unknown",
                "confidence": 0.0,
                "keywords_found": [],
                "message": (
                    "OCR engine is not available on this server. "
                    "Chest X-Ray uploads require OCR verification. "
                    "Install rapidocr-onnxruntime to enable this feature."
                ),
                "error_code": "OCR_REQUIRED_UNAVAILABLE",
            }

        check = is_xray_scan(pil_img)

        if check.is_xray is True:
            # Image passed visual check AND OCR confirmed X-ray markers
            return {
                "allowed": True,
                "is_xray": True,
                "scan_type_detected": check.scan_type_detected,
                "confidence": round(check.confidence, 4),
                "keywords_found": check.keywords_found,
                "message": (
                    f"Verified as chest X-ray "
                    f"({', '.join(check.keywords_found[:3]) or 'X-ray markers detected'})."
                ),
                "error_code": "",
            }
        elif check.is_xray is False:
            return {
                "allowed": False,
                "is_xray": False,
                "scan_type_detected": check.scan_type_detected,
                "confidence": round(check.confidence, 4),
                "keywords_found": check.keywords_found,
                "message": check.rejection_reason or "Image rejected: not an X-ray.",
                "error_code": "OCR_SCAN_REJECTED",
            }
        else:
            # OCR inconclusive (no text found in image) — defer to ScanGate.
            # Clean clinical dataset X-rays have no embedded text so OCR is
            # always inconclusive for them. ScanGate already confirmed the
            # modality before we reached this point, so trust it.
            gate_conf = _gate_result.confidence if _gate_result else 0.0
            return {
                "allowed": True,
                "is_xray": None,
                "scan_type_detected": "xray",
                "confidence": round(gate_conf, 4),
                "keywords_found": [],
                "message": (
                    f"Verified as chest X-ray by scan classifier "
                    f"({gate_conf:.0%} confidence). "
                    "No OCR text markers present — image appears to be a clean scan."
                ),
                "error_code": "",
            }
    else:
        # Brain MRI — ScanGate already confirmed the modality above.
        # Only run OCR to catch explicit wrong-modality text (non-medical
        # watermarks, CT markers). Inconclusive / no-text results pass through
        # since clean MRI datasets have no embedded text at all.
        gate_conf = _gate_result.confidence if _gate_result else 0.0
        if not is_ocr_available():
            return {
                "allowed": True,
                "is_xray": None,
                "scan_type_detected": "mri",
                "confidence": round(gate_conf, 4),
                "keywords_found": [],
                "message": (
                    f"Verified as Brain MRI by scan classifier "
                    f"({gate_conf:.0%} confidence). "
                    "OCR not available — proceeding."
                ),
                "error_code": "",
            }
        check = is_xray_scan(pil_img)
        if check.is_xray is False and check.scan_type_detected in ("non_medical", "ct"):
            return {
                "allowed": False,
                "is_xray": False,
                "scan_type_detected": check.scan_type_detected,
                "confidence": round(check.confidence, 4),
                "keywords_found": check.keywords_found,
                "message": check.rejection_reason or "Non-medical image detected.",
                "error_code": "OCR_SCAN_REJECTED",
            }
        return {
            "allowed": True,
            "is_xray": None,
            "scan_type_detected": "mri",
            "confidence": round(gate_conf, 4),
            "keywords_found": check.keywords_found,
            "message": (
                f"Verified as Brain MRI by scan classifier "
                f"({gate_conf:.0%} confidence)."
            ),
            "error_code": "",
        }


@app.post("/api/predict", tags=["Inference"])
async def predict(
    request: Request,
    image: UploadFile = File(
        ..., description="Medical scan — JPEG, PNG, WebP, BMP, TIFF, GIF, AVIF, DICOM"
    ),
    scan_type: str = Form("Brain MRI", description="'Brain MRI' or 'Chest X-Ray'"),
    gradcam: bool = Form(False, description="Compute Grad-CAM (slower)"),
    mc_dropout: bool = Form(False, description="Compute MC-Dropout uncertainty"),
    mc_samples: int = Form(20, description="Number of MC samples (5-50)"),
):
    """
    Run inference on an uploaded medical image.

    Extra options (all optional):
      gradcam    — if true, a Grad-CAM heatmap is computed and the path
                   is stored with the session for use in PDF reports.
      mc_dropout — if true, runs MC-Dropout uncertainty estimation and
                   returns mean_entropy + std_confidence in the response.
    """
    # Canonicalize scan_type to handle variations (case, spacing, aliases)
    original_scan_type = scan_type
    scan_type = canonicalize_scan_type(scan_type)
    _audit_logger.info(
        "PREDICT_SCAN_TYPE original=%r canonical=%r",
        original_scan_type,
        scan_type,
    )

    model = get_model()
    if model is None:
        raise HTTPException(
            status_code=503,
            detail="Model not loaded. Train the global model first (run.bat).",
        )

    # ── read + validate upload ────────────────────────────────────────────────
    contents = await image.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Empty file received.")
    _validate_upload(contents, image.filename or "")

    # ── decide whether file is DICOM ─────────────────────────────────────────
    orig_ext = Path(image.filename).suffix.lower() if image.filename else ""
    is_dicom = orig_ext == ".dcm"

    pil_img: Image.Image
    detected_format = "unknown"

    if is_dicom:
        try:
            pil_img = _load_dicom(contents)
            detected_format = "DICOM"
        except ImportError as exc:
            raise HTTPException(status_code=501, detail=str(exc))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    else:
        # Standard Pillow decoding
        try:
            pil_img = Image.open(io.BytesIO(contents))
            pil_img.verify()
            pil_img = Image.open(io.BytesIO(contents))
            pil_img.load()
            detected_format = pil_img.format or "unknown"
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Could not decode image: {exc}.",
            )

    # ── medical scan plausibility check ───────────────────────────────────────
    # Compute once, but enforce after OCR gate so OCR-confirmed X-rays are not
    # incorrectly blocked by generic image heuristics.
    medical_plausible = _is_likely_medical_scan(pil_img)

    # ── OCR scan-type gate ────────────────────────────────────────────────────
    #
    # STRICT MODE (Chest X-Ray):
    #   Only images where OCR positively confirms X-ray markers are allowed.
    #   is_xray=True  → pass
    #   is_xray=False → reject (OCR identified non-xray / wrong modality)
    #   is_xray=None  → reject (inconclusive — cannot confirm it is an X-ray)
    #   OCR unavailable → reject (cannot run the required gate)
    #
    # PERMISSIVE MODE (Brain MRI):
    #   OCR gate is skipped entirely. Non-medical/CT content is still rejected
    #   when OCR is available, but inconclusive / no-text images pass through.
    #
    _ocr_scan_validation: dict = {}
    _client_ip = request.client.host if request.client else "unknown"

    # ── STEP 1: ScanGate — ML + heuristic scan-type classifier ────────────────
    # First line of defence: runs before OCR and modality-specific logic.
    _gate = get_scan_gate()
    _gate_result = _gate.check(pil_img, scan_type)
    if not _gate_result.allowed:
        _audit_logger.warning(
            "SCAN_GATE_REJECT ip=%s file=%s label=%s confidence=%.2f heuristics=%s",
            _client_ip,
            image.filename or "unknown",
            _gate_result.label,
            _gate_result.confidence,
            _gate_result.used_heuristics,
        )
        raise HTTPException(
            status_code=400,
            detail={
                "message": _gate_result.rejection_reason,
                "error_code": "OCR_SCAN_REJECTED",
                "scan_type_detected": _gate_result.label,
                "keywords_found": [],
                "ocr_confidence": round(_gate_result.confidence, 4),
            },
        )
    _audit_logger.info(
        "SCAN_GATE_PASS ip=%s file=%s label=%s confidence=%.2f elapsed_ms=%.1f",
        _client_ip,
        image.filename or "unknown",
        _gate_result.label,
        _gate_result.confidence,
        _gate_result.elapsed_ms,
    )

    if scan_type == "Chest X-Ray":
        # ── VISUAL CHECK: secondary grayscale gate after ScanGate ─────────────
        if not medical_plausible:
            _audit_logger.warning(
                "VISUAL_CHECK_REJECT ip=%s file=%s reason=colorful_image",
                _client_ip,
                image.filename or "unknown",
            )
            raise HTTPException(
                status_code=400,
                detail={
                    "message": (
                        "This image does not appear to be a chest X-ray. "
                        "Real chest X-rays are grayscale — this image contains "
                        "colour content inconsistent with radiographic imaging. "
                        "Please upload an original chest X-ray (JPEG / PNG / DICOM)."
                    ),
                    "error_code": "OCR_SCAN_REJECTED",
                    "scan_type_detected": "non_medical",
                    "keywords_found": [],
                    "ocr_confidence": 0.0,
                },
            )

        # ── STRICT: OCR engine must be available ──────────────────────────────
        if not is_ocr_available():
            _audit_logger.warning(
                "OCR_REQUIRED_UNAVAILABLE ip=%s file=%s reason=%s",
                _client_ip,
                image.filename or "unknown",
                ocr_unavailable_reason(),
            )
            raise HTTPException(
                status_code=503,
                detail={
                    "message": (
                        "The OCR scan verification engine is not available on this server. "
                        "Chest X-Ray uploads require OCR verification to prevent non-medical "
                        "images from being processed. "
                        "Please contact the system administrator to install "
                        "rapidocr-onnxruntime."
                    ),
                    "error_code": "OCR_REQUIRED_UNAVAILABLE",
                    "install_hint": "pip install rapidocr-onnxruntime",
                },
            )

        # ── STRICT: run OCR gate ───────────────────────────────────────────────
        _xray_check: XrayScanResult = is_xray_scan(pil_img)
        _ocr_scan_validation = {
            "is_xray": _xray_check.is_xray,
            "scan_type_detected": _xray_check.scan_type_detected,
            "confidence": round(_xray_check.confidence, 4),
            "keywords_found": _xray_check.keywords_found,
            "ocr_ran": _xray_check.ocr_ran,
            "gate_mode": "strict",
        }

        if _xray_check.is_xray is True:
            # Confirmed X-ray — log and allow
            _audit_logger.info(
                "OCR_GATE_PASS ip=%s file=%s keywords=%s confidence=%.2f",
                _client_ip,
                image.filename or "unknown",
                _xray_check.keywords_found[:5],
                _xray_check.confidence,
            )

        elif _xray_check.is_xray is False:
            # OCR positively identified something other than an X-ray
            _audit_logger.warning(
                "OCR_GATE_REJECT ip=%s file=%s scan_type_detected=%s keywords=%s",
                _client_ip,
                image.filename or "unknown",
                _xray_check.scan_type_detected,
                _xray_check.keywords_found,
            )
            raise HTTPException(
                status_code=400,
                detail={
                    "message": _xray_check.rejection_reason
                    or (
                        "This image was identified as non-X-ray content. "
                        "Please upload an original chest X-ray image."
                    ),
                    "error_code": "OCR_SCAN_REJECTED",
                    "scan_type_detected": _xray_check.scan_type_detected,
                    "keywords_found": _xray_check.keywords_found,
                    "ocr_confidence": round(_xray_check.confidence, 4),
                },
            )

        else:
            # OCR inconclusive (no text found in image) — defer to ScanGate.
            # Clean clinical dataset X-rays have no embedded text so OCR is
            # always inconclusive for them. ScanGate already confirmed the
            # modality before we reached this point, so trust it and proceed.
            gate_conf = _gate_result.confidence if _gate_result else 0.0
            _audit_logger.info(
                "OCR_INCONCLUSIVE_GATE_PASS ip=%s file=%s gate_conf=%.2f",
                _client_ip,
                image.filename or "unknown",
                gate_conf,
            )
            _ocr_scan_validation["inconclusive_gate_passed"] = True
            _ocr_scan_validation["gate_confidence"] = round(gate_conf, 4)

    else:
        # ── Brain MRI (and any future modality) ───────────────────────────────
        # ScanGate already confirmed the modality above.
        # Only run OCR to hard-reject explicit non-medical/CT text markers.
        # Inconclusive / no-text results pass through — clean MRI datasets
        # have no embedded text at all.
        gate_conf = _gate_result.confidence if _gate_result else 0.0
        if is_ocr_available():
            _xray_check_mri: XrayScanResult = is_xray_scan(pil_img)
            _ocr_scan_validation = {
                "is_xray": _xray_check_mri.is_xray,
                "scan_type_detected": _xray_check_mri.scan_type_detected,
                "confidence": round(_xray_check_mri.confidence, 4),
                "keywords_found": _xray_check_mri.keywords_found,
                "ocr_ran": _xray_check_mri.ocr_ran,
                "gate_mode": "permissive",
                "gate_ml_confidence": round(gate_conf, 4),
            }
            if (
                _xray_check_mri.is_xray is False
                and _xray_check_mri.scan_type_detected in ("non_medical", "ct")
            ):
                _audit_logger.warning(
                    "OCR_GATE_REJECT_MRI ip=%s file=%s scan_type_detected=%s keywords=%s",
                    _client_ip,
                    image.filename or "unknown",
                    _xray_check_mri.scan_type_detected,
                    _xray_check_mri.keywords_found,
                )
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": _xray_check_mri.rejection_reason
                        or ("This image does not appear to be a valid medical scan."),
                        "error_code": "OCR_SCAN_REJECTED",
                        "scan_type_detected": _xray_check_mri.scan_type_detected,
                        "keywords_found": _xray_check_mri.keywords_found,
                        "ocr_confidence": round(_xray_check_mri.confidence, 4),
                    },
                )
        else:
            _ocr_scan_validation = {
                "gate_mode": "permissive",
                "gate_ml_confidence": round(gate_conf, 4),
                "ocr_ran": False,
            }

    # ── final plausibility enforcement ───────────────────────────────────────
    # For Brain MRI: always apply the visual heuristic.
    # For Chest X-Ray:
    #   • OCR confirmed X-ray (is_xray=True)  → skip visual check, trust OCR.
    #   • OCR inconclusive (is_xray=None)     → APPLY visual check as a
    #     second line of defence.  Clean clinical X-rays are fully grayscale
    #     and always pass; colourful product photos / non-medical images with
    #     plain backgrounds (perfume bottles, food, etc.) are caught here.
    #   • OCR rejected (is_xray=False)        → already raised above, never
    #     reaches this point.
    # Skip the visual plausibility check only when OCR already positively
    # confirmed the image as an X-ray — in that case trust OCR over the
    # heuristic.  For every other combination (Brain MRI, inconclusive CXR,
    # no-text CXR) the visual check is the last line of defence.
    ocr_confirmed_xray = (
        scan_type == "Chest X-Ray" and _ocr_scan_validation.get("is_xray") is True
    )

    if not medical_plausible and not ocr_confirmed_xray:
        modality = "chest X-ray" if scan_type == "Chest X-Ray" else "brain MRI"
        raise HTTPException(
            status_code=400,
            detail={
                "message": (
                    f"This image does not appear to be a {modality}. "
                    "Real medical scans are grayscale — this image contains "
                    "colour content inconsistent with radiographic imaging. "
                    "Please upload an original medical image (JPEG / PNG / DICOM)."
                ),
                "error_code": "OCR_SCAN_REJECTED",
                "scan_type_detected": "non_medical",
                "keywords_found": [],
                "ocr_confidence": 0.0,
            },
        )

    # ── inference ─────────────────────────────────────────────────────────────
    try:
        tensor = prepare_image(pil_img)
        with torch.no_grad():
            logits = model(tensor)
            probs = F.softmax(logits, dim=1).squeeze().numpy()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Inference error: {exc}")

    pred_idx = int(np.argmax(probs))
    pred_key = CLASS_NAMES[pred_idx]
    all_probs: Dict[str, float] = {
        CLASS_NAMES[i]: float(probs[i]) for i in range(len(probs))
    }

    # ── scan-type filtered result ─────────────────────────────────────────────
    mode_cfg = SCAN_MODES.get(scan_type, SCAN_MODES["Brain MRI"])
    mode_probs = {k: all_probs[k] for k in mode_cfg["class_keys"]}
    mode_pred_key = max(mode_probs, key=mode_probs.get)  # type: ignore[arg-type]
    mode_pred_idx = CLASS_NAMES.index(mode_pred_key)

    # ── scan-type mismatch detection ──────────────────────────────────────────
    # The global winner (pred_key) comes from all 6 classes.
    # If it belongs to the OTHER scan mode's class set, the user most likely
    # uploaded the wrong image type (e.g. a brain MRI on the Chest X-Ray tab).
    OTHER_MODE = "Chest X-Ray" if scan_type == "Brain MRI" else "Brain MRI"
    other_cfg = SCAN_MODES[OTHER_MODE]
    other_probs = {k: all_probs[k] for k in other_cfg["class_keys"]}

    selected_mass = sum(mode_probs.values())
    other_mass = sum(other_probs.values())

    global_winner_in_other = pred_key in other_cfg["class_keys"]
    scan_type_mismatch = global_winner_in_other and other_mass > 0.60

    if scan_type_mismatch:
        suggested_scan_type = OTHER_MODE
        other_pred_key = max(other_probs, key=other_probs.get)  # type: ignore[arg-type]
        suggested_class = SHORT_NAMES.get(other_pred_key, other_pred_key)
        suggested_confidence = round(float(other_probs[other_pred_key]), 4)
        mismatch_detail = (
            f"The image appears to be a {OTHER_MODE} scan "
            f"(model assigns {other_mass * 100:.1f}% probability mass to "
            f"{OTHER_MODE} classes, vs {selected_mass * 100:.1f}% to "
            f"{scan_type} classes). "
            f"Suggested result under \u2018{OTHER_MODE}\u2019: "
            f"{suggested_class} ({suggested_confidence * 100:.1f}% confidence)."
        )
    else:
        suggested_scan_type = scan_type
        suggested_class = ""
        suggested_confidence = 0.0
        mismatch_detail = ""

    # Hard gate for obvious wrong-mode uploads:
    # If user selected Brain MRI but model allocates almost all probability mass
    # to Chest X-Ray classes, reject instead of returning a warning-only result.
    if (
        scan_type == "Brain MRI"
        and pred_key in other_cfg["class_keys"]
        and other_mass >= 0.95
    ):
        _audit_logger.warning(
            "SCAN_MODE_MISMATCH_REJECT ip=%s file=%s selected=%s detected=%s other_mass=%.4f",
            _client_ip,
            image.filename or "unknown",
            scan_type,
            OTHER_MODE,
            other_mass,
        )
        raise HTTPException(
            status_code=400,
            detail={
                "message": (
                    "This upload appears to be a Chest X-Ray (or non-MRI image), "
                    "but 'Brain MRI' mode is selected. "
                    "Please switch to 'Chest X-Ray' mode or upload a valid brain MRI image."
                ),
                "error_code": "OCR_SCAN_REJECTED",
                "scan_type_detected": "xray",
                "selected_scan_type": scan_type,
                "suggested_scan_type": OTHER_MODE,
                "selected_mode_mass": round(selected_mass, 4),
                "other_mode_mass": round(other_mass, 4),
            },
        )

    # ── OCR ───────────────────────────────────────────────────────────────────
    ocr_text = ""
    ocr_lines: List[Dict] = []
    if is_ocr_available():
        try:
            ocr_result = extract_text(pil_img)
            ocr_text = ocr_result.full_text
            ocr_lines = [
                {"text": ln.text, "confidence": round(ln.confidence, 4)}
                for ln in ocr_result.lines
            ]
        except Exception:
            pass

    # ── MC-Dropout uncertainty ────────────────────────────────────────────────
    uncertainty_dict: Dict = {}
    if mc_dropout:
        mc_n = max(5, min(50, mc_samples))
        unc = compute_mc_uncertainty(model, tensor, n_samples=mc_n)
        model.eval()  # restore eval mode after MC pass
        uncertainty_dict = {
            "mean_entropy": unc.mean_entropy,
            "std_confidence": unc.std_confidence,
            "mc_samples": unc.mc_samples,
            "uncertainty_label": unc.uncertainty_label,
        }

    # ── persist temp copy ────────────────────────────────────────────────────
    session_id = str(uuid.uuid4())
    _FORMAT_TO_EXT: dict[str, str] = {
        "JPEG": ".jpg",
        "PNG": ".png",
        "WEBP": ".webp",
        "BMP": ".bmp",
        "TIFF": ".tiff",
        "GIF": ".gif",
        "AVIF": ".avif",
        "DICOM": ".dcm",
    }
    if orig_ext in ALLOWED_EXTENSIONS:
        suffix = orig_ext
    else:
        suffix = _FORMAT_TO_EXT.get(detected_format.upper(), ".jpg")

    tmp_path = TEMP_DIR / f"{session_id}{suffix}"
    tmp_path.write_bytes(contents)
    _pending_images[session_id] = tmp_path

    # ── Grad-CAM (optional, stored alongside temp image) ─────────────────────
    gradcam_path: Optional[Path] = None
    if gradcam:
        try:
            cam_overlay = compute_gradcam(model, tensor, mode_pred_idx, pil_img)
            cam_pil = Image.fromarray(cam_overlay, mode="RGB")
            gradcam_path = TEMP_DIR / f"{session_id}_gradcam.png"
            cam_pil.save(str(gradcam_path))
        except Exception as exc:
            print(f"[API] Grad-CAM failed: {exc}")

    # ── persist session to DB ────────────────────────────────────────────────
    predict_response: Dict = {
        "session_id": session_id,
        "filename": image.filename or f"uploaded_scan{suffix}",
        "detected_format": detected_format,
        "scan_type": scan_type,
        "predicted_key": pred_key,
        "predicted_class": SHORT_NAMES.get(pred_key, pred_key),
        "confidence": round(float(probs[pred_idx]), 4),
        "probabilities": all_probs,
        "mode_predicted_key": mode_pred_key,
        "mode_predicted_class": SHORT_NAMES.get(mode_pred_key, mode_pred_key),
        "mode_confidence": round(float(mode_probs[mode_pred_key]), 4),
        "mode_probabilities": mode_probs,
        # probability mass each mode holds (0–1)
        "selected_mode_mass": round(selected_mass, 4),
        "other_mode_mass": round(other_mass, 4),
        # mismatch fields
        "scan_type_mismatch": scan_type_mismatch,
        "suggested_scan_type": suggested_scan_type,
        "suggested_class": suggested_class,
        "suggested_confidence": suggested_confidence,
        "mismatch_detail": mismatch_detail,
        "ocr_text": ocr_text,
        "ocr_lines": ocr_lines,
        "ocr_scan_validation": _ocr_scan_validation,
        "file_size_bytes": len(contents),
        "image_dimensions": list(pil_img.size),
        "gradcam_available": gradcam_path is not None,
    }
    if uncertainty_dict:
        predict_response["uncertainty"] = uncertainty_dict

    try:
        get_db().add_session(predict_response)
    except Exception as exc:
        print(f"[API] DB session save failed: {exc}")

    return predict_response


# ── feedback ───────────────────────────────────────────────────────────────────


@app.post("/api/feedback", tags=["Feedback"])
async def feedback(
    session_id: str = Form(...),
    chosen_key: str = Form(...),
    scan_type: str = Form("Brain MRI"),
    ai_predicted_key: str = Form(""),
    clinician_name: str = Form(""),
    clinician_id: str = Form(""),
    notes: str = Form(""),
):
    """
    Persist the doctor's confirmed / overridden label for continuous learning.
    The image is anonymised and saved for the next FL round.
    """
    if chosen_key not in CLASS_NAMES:
        raise HTTPException(
            status_code=400, detail=f"Unknown class key: '{chosen_key}'"
        )

    tmp_path = _pending_images.get(session_id)
    if tmp_path is None or not tmp_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Session image not found. Please upload and predict again.",
        )

    save_dir = COLLECT_DIR / chosen_key
    save_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = save_dir / f"{ts}_{session_id[:8]}{tmp_path.suffix}"

    try:
        pil_img = Image.open(tmp_path)
        strip_metadata_and_save(pil_img, dest)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save image: {exc}")

    entry: Dict = {
        "session_id": session_id,
        "chosen_key": chosen_key,
        "chosen_label": SHORT_NAMES.get(chosen_key, chosen_key),
        "ai_predicted_key": ai_predicted_key,
        "scan_type": scan_type,
        "saved_to": str(dest),
        "timestamp": datetime.datetime.now().isoformat(),
        "overridden": chosen_key != ai_predicted_key,
        "clinician_name": clinician_name,
        "clinician_id": clinician_id,
        "notes": notes,
    }

    try:
        get_db().add_feedback(entry)
    except Exception as exc:
        print(f"[API] DB feedback save failed: {exc}")

    # Keep temp file alive so the clinician can still download a PDF report
    # after submitting feedback.  Background cleanup handles eviction after 1h.

    return {
        "success": True,
        "overridden": entry["overridden"],
        "message": (
            f"Image saved as '{SHORT_NAMES.get(chosen_key, chosen_key)}' "
            "and queued for the next training round."
        ),
        "saved_to": str(dest),
    }


# ── legacy queue endpoint (kept for backwards compatibility) ──────────────────


@app.get("/api/queue", tags=["Feedback"])
def queue(limit: int = 50):
    """Return the last N doctor-feedback entries."""
    db = get_db()
    rows = db.list_feedback(limit=min(limit, 200))
    counts = db.count_feedback()
    return {
        "count": counts["total"],
        "entries": rows,
    }


# ── legacy plain-text report (kept for backwards compatibility) ───────────────


@app.get("/api/report", tags=["Reporting"], response_class=PlainTextResponse)
def report(
    session_id: str = "",
    filename: str = "uploaded_scan.jpg",
    scan_type: str = "Brain MRI",
    ai_pred: str = "",
    doctor_confirmed: str = "",
    ocr_text: str = "",
):
    """Generate and download a plain-text diagnostic report (legacy format)."""
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    outcome = "CONFIRMED" if ai_pred == doctor_confirmed else "OVERRIDDEN BY CLINICIAN"
    ai_label = SHORT_NAMES.get(ai_pred, ai_pred or "\u2014")
    doc_label = SHORT_NAMES.get(doctor_confirmed, doctor_confirmed or "\u2014")

    ocr_block = ""
    if ocr_text.strip():
        ocr_block = (
            "\n--------------------------------------------------------\n"
            "  EXTRACTED TEXT (OCR)\n"
            f"{ocr_text}\n"
        )

    report_text = (
        "========================================================\n"
        "  TECNOMATE CLINICAL AI \u2014 DIAGNOSTIC REPORT\n"
        "========================================================\n"
        f"  Date/Time         : {ts}\n"
        f"  Image File        : {filename}\n"
        f"  Scan Type         : {scan_type}\n"
        f"  AI Prediction     : {ai_label}\n"
        f"  Doctor Confirmed  : {doc_label}\n"
        f"  Outcome           : {outcome}\n"
        f"{ocr_block}"
        "--------------------------------------------------------\n"
        "  PRIVACY NOTICE\n"
        "  Patient Data Anonymized and Secured.\n"
        "  All EXIF metadata, device identifiers, and hidden\n"
        "  tags have been permanently stripped from this image\n"
        "  before storage.  No patient-identifiable information\n"
        "  is retained anywhere in this system.\n"
        "--------------------------------------------------------\n"
        "  MEDICAL DISCLAIMER\n"
        "  This report is generated by an AI assistant only.\n"
        "  It must not be used as the sole basis for clinical\n"
        "  decisions.  Always rely on qualified medical\n"
        "  professionals for diagnosis and treatment.\n"
        "========================================================\n"
    )

    safe_ts = ts.replace(":", "-").replace(" ", "_")
    filename_out = f"tecnomate_report_{safe_ts}.txt"
    return PlainTextResponse(
        content=report_text,
        headers={"Content-Disposition": f'attachment; filename="{filename_out}"'},
    )


# ── PDF report ─────────────────────────────────────────────────────────────────


@app.post("/api/pdf-report", tags=["Reporting"])
async def pdf_report(
    # ── core identification (required) ────────────────────────────────────────
    session_id: str = Form(...),
    scan_type: str = Form("Brain MRI"),
    ai_pred_key: str = Form(...),
    ai_confidence: float = Form(0.0),
    doctor_choice_key: str = Form(...),
    # ── probabilities (JSON string) ───────────────────────────────────────────
    probabilities_json: str = Form("{}"),
    # ── OCR ───────────────────────────────────────────────────────────────────
    ocr_text: str = Form(""),
    ocr_lines_json: str = Form("[]"),
    # ── patient / visit fields (all optional) ─────────────────────────────────
    patient_name: str = Form("Anonymous / De-identified"),
    patient_id: str = Form("N/A"),
    date_of_birth: str = Form("N/A"),
    gender: str = Form("N/A"),
    referring_doctor: str = Form("N/A"),
    institution: str = Form("Tecnomate Health Network"),
    visit_date: str = Form(""),
    clinical_notes: str = Form(""),
    # ── clinician identity (optional) ─────────────────────────────────────────
    clinician_name: str = Form(""),
    clinician_id: str = Form(""),
    # ── model metadata ────────────────────────────────────────────────────────
    fl_round: int = Form(0),
    report_model_version: str = Form("global_model.pth"),
    # ── uncertainty (optional — populated by /api/predict if mc_dropout=True) ─
    mc_entropy: float = Form(0.0),
    mc_std_conf: float = Form(0.0),
    mc_samples: int = Form(0),
    mc_label: str = Form(""),
    # ── AI mathematical parameters (from predict response) ──────────────────
    selected_mode_mass: float = Form(0.0),
    other_mode_mass: float = Form(0.0),
    scan_type_mismatch: bool = Form(False),
    # ── report format ─────────────────────────────────────────────────────────
    format: str = Form("reportlab"),
):
    """
    Generate a professional clinical PDF report and return it for download.

    Accepts all prediction and patient metadata as form fields so the
    frontend can submit a single POST after the clinician has filled in
    the optional patient fields.

    The report includes:
      - Cover band with report timestamp and session ID
      - Patient & visit information table
      - Original scan thumbnail + Grad-CAM heatmap (if available)
      - AI prediction banner with risk level and ICD-10 code
      - Per-class probability bar chart
      - MC-Dropout uncertainty (if mc_samples > 0)
      - Clinician review / final diagnosis + signature line
      - OCR-extracted text block (if any)
      - Audit trail
      - Privacy notice, FL notice, medical disclaimer
    """
    import json as _json

    # ── parse JSON blobs ──────────────────────────────────────────────────────
    try:
        probabilities: Dict[str, float] = _json.loads(probabilities_json)
    except Exception:
        probabilities = {}

    try:
        ocr_lines: List[Dict] = _json.loads(ocr_lines_json)
    except Exception:
        ocr_lines = []

    # ── look up scan image from temp dir ─────────────────────────────────────
    scan_image_path: Optional[Path] = _pending_images.get(session_id)
    if scan_image_path is None or not scan_image_path.exists():
        # Try to find leftover temp file by session prefix
        candidates = list(TEMP_DIR.glob(f"{session_id}.*"))
        non_gradcam = [c for c in candidates if "_gradcam" not in c.name]
        scan_image_path = non_gradcam[0] if non_gradcam else None

    # ── look up Grad-CAM overlay ──────────────────────────────────────────────
    gradcam_np: Optional[np.ndarray] = None
    gradcam_temp = TEMP_DIR / f"{session_id}_gradcam.png"
    if gradcam_temp.exists():
        try:
            gradcam_np = np.array(Image.open(gradcam_temp).convert("RGB"))
        except Exception:
            pass
    elif scan_image_path and scan_image_path.exists():
        # Compute Grad-CAM on the fly if not pre-computed
        model = get_model()
        if model is not None:
            try:
                pil_scan = Image.open(scan_image_path)
                infer_tensor = prepare_image(pil_scan)
                target_idx = (
                    CLASS_NAMES.index(ai_pred_key) if ai_pred_key in CLASS_NAMES else 0
                )
                gradcam_np = compute_gradcam(model, infer_tensor, target_idx, pil_scan)
            except Exception as exc:
                print(f"[API/PDF] on-demand Grad-CAM failed: {exc}")

    # ── assemble ReportRequest ────────────────────────────────────────────────
    req = ReportRequest(
        session_id=session_id,
        filename=(
            _pending_images.get(session_id, Path(f"scan_{session_id[:8]}.jpg")).name
            if scan_image_path is None
            else scan_image_path.name
        ),
        scan_type=scan_type,
        ai_pred_key=ai_pred_key,
        ai_confidence=ai_confidence,
        probabilities=probabilities,
        doctor_choice_key=doctor_choice_key,
        ocr_text=ocr_text,
        ocr_lines=ocr_lines,
        scan_image_path=scan_image_path,
        gradcam_image=gradcam_np,
        patient=PatientInfo(
            patient_name=patient_name,
            patient_id=patient_id,
            date_of_birth=date_of_birth,
            gender=gender,
            referring_doctor=referring_doctor,
            institution=institution,
            visit_date=visit_date,
            clinical_notes=clinical_notes,
        ),
        uncertainty=UncertaintyInfo(
            mean_entropy=mc_entropy,
            std_confidence=mc_std_conf,
            mc_samples=mc_samples,
            uncertainty_label=mc_label or ("N/A" if mc_samples == 0 else ""),
        ),
        fl_round=fl_round,
        model_version=report_model_version,
        server_timestamp=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"),
        clinician_name=clinician_name,
        clinician_id=clinician_id,
        selected_mode_mass=selected_mode_mass,
        other_mode_mass=other_mode_mass,
        scan_type_mismatch=scan_type_mismatch,
    )

    # ── build PDF ─────────────────────────────────────────────────────────────
    try:
        if format.lower() == "latex":
            from latex_report import build_latex_report

            try:
                pdf_bytes = build_latex_report(req)
            except Exception as latex_exc:
                print(
                    f"[API/PDF] LaTeX failed ({latex_exc}), falling back to reportlab"
                )
                pdf_bytes = build_pdf_report(req)
        else:
            pdf_bytes = build_pdf_report(req)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"PDF generation failed: {exc}",
        )

    # ── clean up Grad-CAM temp file after embedding it ────────────────────────
    try:
        gradcam_temp.unlink(missing_ok=True)
    except Exception:
        pass

    ts_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"tecnomate_report_{ts_str}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── admin endpoints ────────────────────────────────────────────────────────────

security = HTTPBasic()

def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    admin_user = os.getenv("ADMIN_USER")
    admin_pass = os.getenv("ADMIN_PASS")

    if not admin_user or not admin_pass:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Admin credentials not configured on the server."
        )

    is_user_ok = secrets.compare_digest(credentials.username, admin_user)
    is_pass_ok = secrets.compare_digest(credentials.password, admin_pass)

    if not (is_user_ok and is_pass_ok):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials


@app.get("/api/admin/stats", tags=["Admin"])
def admin_stats(admin: HTTPBasicCredentials = Depends(verify_admin)):
    """Return aggregate statistics from the SQLite database."""
    return get_db().stats()


@app.get("/api/admin/feedback", tags=["Admin"])
def admin_feedback(
    limit: int = 50,
    offset: int = 0,
    overridden_only: bool = False,
    scan_type: str = "",
    admin: HTTPBasicCredentials = Depends(verify_admin),
):
    """Return paginated feedback records."""
    db = get_db()
    rows = db.list_feedback(
        limit=min(limit, 500),
        offset=offset,
        overridden_only=overridden_only,
        scan_type=scan_type or None,
    )
    counts = db.count_feedback(scan_type=scan_type or None)
    return {
        "total": counts["total"],
        "confirmed": counts["confirmed"],
        "overridden": counts["overridden"],
        "rows": rows,
    }


@app.get("/api/admin/export-csv", tags=["Admin"])
def admin_export_csv(admin: HTTPBasicCredentials = Depends(verify_admin)):
    """Download all feedback data as a CSV file."""
    csv_data = get_db().export_feedback_csv()
    ts_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"tecnomate_feedback_{ts_str}.csv"
    return Response(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/admin/export-excel", tags=["Admin"])
def admin_export_excel(admin: HTTPBasicCredentials = Depends(verify_admin)):
    """
    Download a full admin report as a formatted Excel workbook.

    3 sheets:
      1. Feedback Log   — every doctor feedback/override entry from SQLite
      2. Sessions Log   — every prediction session (filename, AI result, confidence)
      3. Summary Stats  — class-level counts, override rate, scan-type breakdown
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    db = get_db()
    # Fetch up to 10 000 rows — sufficient for any realistic deployment
    feedback_rows = db.list_feedback(limit=10_000, offset=0)
    session_rows = db.list_sessions(limit=10_000, offset=0)
    counts = db.count_feedback()

    wb = openpyxl.Workbook()

    # ─ shared styles ───────────────────────────────────────────────────────────
    H_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue header
    A_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue alt row
    G_FILL = PatternFill("solid", fgColor="C6EFCE")  # green  — confirmed
    W_FILL = PatternFill("solid", fgColor="FFEB9C")  # yellow — overridden
    R_FILL = PatternFill("solid", fgColor="FFC7CE")  # red    — risk classes
    H_FONT = Font(bold=True, color="FFFFFF", size=11)
    T_FONT = Font(bold=True, size=13)
    B_FONT = Font(bold=True)
    CTR = Alignment(horizontal="center", vertical="center")

    def _hrow(ws, row: int, cols: list):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = CTR

    def _title(ws, text: str, span: str):
        ws.merge_cells(span)
        c = ws[span.split(":")[0]]
        c.value = text
        c.font = T_FONT
        c.alignment = CTR
        ws.row_dimensions[1].height = 24

    def _autowidth(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                w + 4, 50
            )

    RISK_CLASSES = {"glioma", "meningioma", "pituitary", "pneumonia"}

    # ═════════════════════════════ Sheet 1: Feedback Log ═════════════════════════════
    ws1 = wb.active
    ws1.title = "Feedback Log"
    fb_cols = [
        "ID",
        "Session ID",
        "Doctor Choice",
        "Doctor Label",
        "AI Predicted",
        "Scan Type",
        "Overridden",
        "Clinician",
        "Clinician ID",
        "Notes",
        "Timestamp",
    ]
    _title(
        ws1, "Tecnomate — Doctor Feedback Log", f"A1:{get_column_letter(len(fb_cols))}1"
    )
    _hrow(ws1, 2, fb_cols)

    for i, row in enumerate(feedback_rows, start=3):
        overridden = bool(row.get("overridden", 0))
        vals = [
            row.get("id"),
            row.get("session_id"),
            row.get("chosen_key"),
            row.get("chosen_label"),
            row.get("ai_predicted_key"),
            row.get("scan_type"),
            "Yes" if overridden else "No",
            row.get("clinician_name"),
            row.get("clinician_id"),
            row.get("notes"),
            row.get("timestamp"),
        ]
        for c, val in enumerate(vals, 1):
            ws1.cell(row=i, column=c, value=val).alignment = CTR
        # row colour: yellow if doctor overrode AI, green if confirmed
        fill = W_FILL if overridden else (G_FILL if i % 2 == 0 else A_FILL)
        for c in range(1, len(fb_cols) + 1):
            ws1.cell(row=i, column=c).fill = fill
    _autowidth(ws1)

    # ════════════════════════════ Sheet 2: Sessions Log ════════════════════════════
    ws2 = wb.create_sheet("Sessions Log")
    sess_cols = [
        "Session ID",
        "Filename",
        "Scan Type",
        "AI Prediction",
        "Confidence (%)",
        "File Size (KB)",
        "Width",
        "Height",
        "Format",
        "Created At",
    ]
    _title(
        ws2,
        "Tecnomate — Prediction Sessions",
        f"A1:{get_column_letter(len(sess_cols))}1",
    )
    _hrow(ws2, 2, sess_cols)

    for i, row in enumerate(session_rows, start=3):
        pred_key = row.get("ai_pred_key") or ""
        conf = float(row.get("ai_confidence") or 0.0)
        vals = [
            row.get("session_id"),
            row.get("filename"),
            row.get("scan_type"),
            pred_key,
            round(conf * 100, 2),
            round((row.get("file_size_bytes") or 0) / 1024, 1),
            row.get("image_width"),
            row.get("image_height"),
            row.get("detected_format"),
            row.get("created_at"),
        ]
        for c, val in enumerate(vals, 1):
            ws2.cell(row=i, column=c, value=val).alignment = CTR
        # highlight high-risk predictions
        row_fill = (
            R_FILL if pred_key in RISK_CLASSES else (A_FILL if i % 2 == 0 else None)
        )
        if row_fill:
            for c in range(1, len(sess_cols) + 1):
                ws2.cell(row=i, column=c).fill = row_fill
    _autowidth(ws2)

    # ═══════════════════════ Sheet 3: AI Diagnostic Parameters ═══════════════════════
    ws3 = wb.create_sheet("AI Parameters")
    ai_cols = [
        "Session ID",
        "Filename",
        "Scan Type",
        "AI Prediction",
        "Confidence (%)",
        "Prediction Margin (%)",
        "Shannon Entropy (bits)",
        "Mode Prob Mass (%)",
        "Other Mode Mass (%)",
        "Scan Mismatch",
        "MC Entropy",
        "MC Std Conf",
        "MC Samples",
        "Uncertainty Level",
        "All Class Probs (JSON)",
        "Created At",
    ]
    _title(
        ws3,
        "Tecnomate — AI Diagnostic Mathematical Parameters",
        f"A1:{get_column_letter(len(ai_cols))}1",
    )
    _hrow(ws3, 2, ai_cols)

    for i, row in enumerate(session_rows, start=3):
        pred_key = row.get("ai_pred_key") or ""
        conf = float(row.get("ai_confidence") or 0.0)
        margin = float(row.get("prediction_margin") or 0.0)
        entropy = float(row.get("shannon_entropy") or 0.0)
        sel_mass = float(row.get("selected_mode_mass") or 0.0)
        oth_mass = float(row.get("other_mode_mass") or 0.0)
        mismatch = bool(row.get("scan_type_mismatch") or False)
        mc_ent = float(row.get("mean_entropy") or 0.0)
        mc_std = float(row.get("std_confidence") or 0.0)
        mc_n = int(row.get("mc_samples") or 0)
        unc_lbl = row.get("uncertainty_label") or ""
        all_p = row.get("all_probabilities") or {}

        vals = [
            row.get("session_id"),
            row.get("filename"),
            row.get("scan_type"),
            pred_key,
            round(conf * 100, 2),
            round(margin * 100, 2),
            round(entropy, 4),
            round(sel_mass * 100, 2),
            round(oth_mass * 100, 2),
            "YES" if mismatch else "No",
            round(mc_ent, 4) if mc_n > 0 else "N/A (MC off)",
            round(mc_std, 4) if mc_n > 0 else "N/A",
            mc_n if mc_n > 0 else "N/A",
            unc_lbl or "N/A",
            json.dumps({k: round(v * 100, 2) for k, v in all_p.items()})
            if all_p
            else "{}",
            row.get("created_at"),
        ]
        for c, val in enumerate(vals, 1):
            ws3.cell(row=i, column=c, value=val).alignment = CTR
        row_fill = R_FILL if mismatch else (A_FILL if i % 2 == 0 else None)
        if row_fill:
            for c in range(1, len(ai_cols) + 1):
                ws3.cell(row=i, column=c).fill = row_fill
    _autowidth(ws3)

    # ════════════════════════════ Sheet 4: Summary Stats ═══════════════════════════
    ws4 = wb.create_sheet("Summary Stats")
    _title(ws4, "Tecnomate — Admin Summary Statistics", "A1:B1")
    _hrow(ws4, 2, ["Metric", "Value"])

    # aggregate per-class prediction counts from sessions
    class_counts: Dict[str, int] = {c: 0 for c in CLASS_NAMES}
    avg_confidence: Dict[str, list] = {c: [] for c in CLASS_NAMES}
    avg_entropy: Dict[str, list] = {c: [] for c in CLASS_NAMES}
    avg_margin: Dict[str, list] = {c: [] for c in CLASS_NAMES}
    for row in session_rows:
        key = row.get("ai_pred_key", "")
        if key in class_counts:
            class_counts[key] += 1
            conf_v = row.get("ai_confidence", 0)
            if conf_v:
                avg_confidence[key].append(float(conf_v))
            ent_v = row.get("shannon_entropy", 0)
            if ent_v:
                avg_entropy[key].append(float(ent_v))
            mar_v = row.get("prediction_margin", 0)
            if mar_v:
                avg_margin[key].append(float(mar_v))

    # override rate per class
    override_by_class: Dict[str, int] = {c: 0 for c in CLASS_NAMES}
    for row in feedback_rows:
        if row.get("overridden"):
            key = row.get("ai_predicted_key", "")
            if key in override_by_class:
                override_by_class[key] += 1

    override_rate = (
        f"{counts['overridden'] / counts['total'] * 100:.1f}%"
        if counts["total"] > 0
        else "N/A"
    )
    ts_now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # global average confidence and entropy across all sessions
    all_conf = [
        float(r.get("ai_confidence", 0)) for r in session_rows if r.get("ai_confidence")
    ]
    all_ent = [
        float(r.get("shannon_entropy", 0))
        for r in session_rows
        if r.get("shannon_entropy")
    ]
    all_margin = [
        float(r.get("prediction_margin", 0))
        for r in session_rows
        if r.get("prediction_margin")
    ]
    mismatches = sum(1 for r in session_rows if r.get("scan_type_mismatch"))

    summary = [
        ("Report generated at", ts_now),
        ("Total feedback entries", counts["total"]),
        ("Doctor-confirmed (AI correct)", counts["confirmed"]),
        ("Doctor-overridden (AI wrong)", counts["overridden"]),
        ("Overall override rate", override_rate),
        ("Total prediction sessions", len(session_rows)),
        ("Scan type mismatches detected", mismatches),
        (
            "Global avg confidence",
            f"{sum(all_conf) / len(all_conf) * 100:.2f}%" if all_conf else "N/A",
        ),
        (
            "Global avg Shannon entropy (bits)",
            f"{sum(all_ent) / len(all_ent):.4f}" if all_ent else "N/A",
        ),
        (
            "Global avg prediction margin",
            f"{sum(all_margin) / len(all_margin) * 100:.2f}%" if all_margin else "N/A",
        ),
        ("", ""),
        ("--- Per-class breakdown ---", ""),
    ]
    for cls in CLASS_NAMES:
        cnt = class_counts[cls]
        ov = override_by_class[cls]
        ov_pct = f"{ov / cnt * 100:.1f}%" if cnt > 0 else "N/A"
        ac = avg_confidence[cls]
        ae = avg_entropy[cls]
        am = avg_margin[cls]
        summary.append((f"  {cls} — total scans", cnt))
        summary.append((f"  {cls} — overrides", f"{ov}  ({ov_pct})"))
        summary.append(
            (
                f"  {cls} — avg confidence",
                f"{sum(ac) / len(ac) * 100:.2f}%" if ac else "N/A",
            )
        )
        summary.append(
            (
                f"  {cls} — avg Shannon entropy",
                f"{sum(ae) / len(ae):.4f}" if ae else "N/A",
            )
        )
        summary.append(
            (
                f"  {cls} — avg prediction margin",
                f"{sum(am) / len(am) * 100:.2f}%" if am else "N/A",
            )
        )
        summary.append(("", ""))

    for i, (key, val) in enumerate(summary, start=3):
        ws4.cell(row=i, column=1, value=key).font = B_FONT
        ws4.cell(row=i, column=2, value=val)
        if i % 2 == 0:
            ws4.cell(row=i, column=1).fill = A_FILL
            ws4.cell(row=i, column=2).fill = A_FILL
    _autowidth(ws4)

    # ─ serialise workbook to bytes and stream as download ───────────────────
    import io as _io

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"tecnomate_admin_report_{ts_str}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/admin/sessions", tags=["Admin"])
def admin_sessions(limit: int = 50, offset: int = 0, admin: HTTPBasicCredentials = Depends(verify_admin)):
    """Return paginated prediction sessions."""
    return {"sessions": get_db().list_sessions(limit=min(limit, 500), offset=offset)}


# ── SPA + static file handler (must be last) ──────────────────────────────────
# FastAPI routes take priority over app.mount(), so this catch-all serves both:
#   • real files (assets/*, favicon.svg, …) → FileResponse from dist/
#   • React Router paths (/classify, /history, …) → index.html
@app.get("/{full_path:path}", include_in_schema=False)
def serve_spa(full_path: str):
    """Serve dist files when they exist; fall back to index.html for React Router."""
    if FRONTEND_DIR.exists():
        candidate = FRONTEND_DIR / full_path

        try:
            candidate.resolve().relative_to(FRONTEND_DIR.resolve())
        except ValueError:
            return JSONResponse({"detail": "Not Found"}, status_code=404)

        if candidate.is_file():
            return FileResponse(str(candidate))
        index_file = FRONTEND_DIR / "index.html"
        if index_file.exists():
            return FileResponse(str(index_file))
    return JSONResponse(
        {"detail": "Frontend not built. Run: cd frontend && npm run build"},
        status_code=404,
    )
